from __future__ import annotations

import json
import math
import re
import shutil
import sqlite3
import sys
import urllib.parse
import urllib.request
from datetime import date
from pathlib import Path
from types import SimpleNamespace
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "raw"
CACHE_DIR = ROOT / "data" / "cache"
TALDAU_CACHE = CACHE_DIR / "taldau_production.json"
DASHBOARD_DIR = ROOT / "dashboard"
OUTPUT_JSON = DASHBOARD_DIR / "data" / "bns.json"
STOCKS_DB = ROOT / "data" / "stocks.sqlite3"
# Fallback is allowed only for data that this project previously received
# from official online APIs (currently Taldau 701608). It is not a permission
# to use manually edited local Excel files as a source.
OFFICIAL_CACHE_ALLOWED = True

sys.path.insert(0, str(ROOT))
from bns_agent.config import SOURCE_PAGE as BNS_PRICE_SOURCE_PAGE  # noqa: E402
from bns_agent.parser import _observation_date, parse_snapshot, product_key  # noqa: E402
from bns_agent.stocks import REGIONS as STOCK_REGIONS  # noqa: E402


SZPT_ORDER = [
    "рис шлифованный",
    "крупа гречневая",
    "мука пшеничная первого сорта",
    "хлеб пшеничный из муки первого сорта",
    "рожки",
    "говядина с костями",
    "говядина бескостная",
    "конина, включая бескостную",
    "баранина, включая бескостную",
    "мясо птицы",
    "мясной фарш",
    "рыба свежая, охлажденная, мороженая (лещ, карась, судак, карп, сазан)",
    "молоко (пастеризованное, ультрапастеризованное, стерилизованное от 2,2% до 6% жирности)",
    "сыр твердый, полутвердый",
    "творог 5-9% жирности",
    "сметана",
    "кефир 2-3% жирности",
    "яйца, 1 категории",
    "масло подсолнечное",
    "масло сливочное",
    "яблоки",
    "капуста белокочанная",
    "огурцы",
    "помидоры",
    "лук репчатый",
    "морковь",
    "картофель",
    "сахар-песок",
    "соль, кроме экстра",
    "чай черный",
]

PRICE_ALIASES = {
    "рис шлифованный": "рис шлифованный",
    "крупа гречневая": "крупа гречневая ядрица",
    "мука пшеничная первого сорта": "мука пшеничная первого сорта",
    "хлеб пшеничный из муки первого сорта": "хлеб пшеничный из муки первого сорта",
    "рожки": "рожки",
    "говядина с костями": "говядина лопаточно-грудная часть с костями",
    "говядина бескостная": "говядина тазобедренная часть бескостная",
    "конина, включая бескостную": "конина с костями",
    "баранина, включая бескостную": "баранина с костями",
    "мясо птицы": "куры",
    "мясной фарш": "мясной фарш",
    "рыба свежая, охлажденная, мороженая (лещ, карась, судак, карп, сазан)": "рыба свежая",
    "молоко (пастеризованное, ультрапастеризованное, стерилизованное от 2,2% до 6% жирности)": "молоко пастеризованное",
    "сыр твердый, полутвердый": "сыр сычужный твердый",
    "творог 5-9% жирности": "творог",
    "сметана": "сметана",
    "кефир 2-3% жирности": "кефир",
    "яйца, 1 категории": "яйца",
    "масло подсолнечное": "масло подсолнечное",
    "масло сливочное": "масло сливочное несоленое",
    "яблоки": "яблоки",
    "капуста белокочанная": "капуста белокочанная",
    "огурцы": "огурцы свежие",
    "помидоры": "помидоры свежие",
    "лук репчатый": "лук репчатый",
    "морковь": "морковь",
    "картофель": "картофель",
    "сахар-песок": "сахар-песок",
    "соль, кроме экстра": "соль",
    "чай черный": "чай черный",
}

PRODUCTION_ROWS: dict[str, list[str]] = {
    "рис шлифованный": ["рис полуобрушенный или полностью обрушенный"],
    "крупа гречневая": ["крупа, мука грубого помола"],
    "мука пшеничная первого сорта": ["мука из культур зерновых"],
    "хлеб пшеничный из муки первого сорта": ["хлеб свежий"],
    "рожки": ["макароны, лапша, кускус"],
    "говядина с костями": ["мясо скота крупного рогатого, свиней, овец, коз, лошадей"],
    "говядина бескостная": ["мясо скота крупного рогатого, свиней, овец, коз, лошадей"],
    "конина, включая бескостную": ["мясо скота крупного рогатого, свиней, овец, коз, лошадей"],
    "баранина, включая бескостную": ["мясо скота крупного рогатого, свиней, овец, коз, лошадей"],
    "мясо птицы": ["мясо птицы домашней, свежее или охлажденное", "куры (включая цыплят)"],
    "мясной фарш": ["полуфабрикаты готовые из мяса"],
    "рыба свежая, охлажденная, мороженая (лещ, карась, судак, карп, сазан)": ["рыба, свежая, охлажденная или мороженая"],
    "молоко (пастеризованное, ультрапастеризованное, стерилизованное от 2,2% до 6% жирности)": ["молоко обработанное жидкое и сливки"],
    "сыр твердый, полутвердый": ["сыры твердые"],
    "творог 5-9% жирности": ["творог нежирный", "творог жирный"],
    "сметана": ["йогурт, молоко и сливки ферментированные"],
    "кефир 2-3% жирности": ["йогурт, молоко и сливки ферментированные"],
    "масло подсолнечное": ["масло подсолнечное нерафинированное", "масло подсолнечное и его фракции"],
    "масло сливочное": ["масло сливочное и спреды"],
    "сахар-песок": ["сахар-сырец или сахар рафинированный"],
    "соль, кроме экстра": ["соль и хлорид натрия чистый"],
    "чай черный": ["чай и кофе переработанные"],
}


OFFICIAL_TRADE_FILE_2025 = RAW_DIR / "foreign_trade" / "converted" / "tab_26_12_2025.xlsx"
OFFICIAL_TRADE_CURRENT_DIR = RAW_DIR / "foreign_trade" / "current"
OFFICIAL_TRADE_CURRENT_META = OFFICIAL_TRADE_CURRENT_DIR / "latest_2026.json"
OFFICIAL_TRADE_SOURCE_2025 = (
    "БНС: Экспорт и импорт товаров РК по 4,6,10 знакам ТН ВЭД ЕАЭС "
    "(январь-декабрь 2025г.)"
)
OFFICIAL_TRADE_URL_2025 = (
    "https://stat.gov.kz/ru/industries/economy/foreign-market/spreadsheets/"
    "?year=2025&name=40113&type=spreadsheets"
)
OFFICIAL_TRADE_URL_CURRENT = (
    "https://stat.gov.kz/ru/industries/economy/foreign-market/spreadsheets/"
    "?year=2026&name=40113&type=spreadsheets"
)
OFFICIAL_CROP_HARVEST_URLS = {
    2024: "https://stat.gov.kz/api/iblock/element/301855/file/ru/",
    2025: "https://stat.gov.kz/api/iblock/element/474819/file/ru/",
}
OFFICIAL_LIVESTOCK_PAGE_URL = (
    "https://stat.gov.kz/ru/industries/business-statistics/"
    "stat-forrest-village-hunt-fish/spreadsheets/?name=18612&type=spreadsheets"
)
OFFICIAL_LIVESTOCK_URLS = {
    "current": "https://stat.gov.kz/api/iblock/element/498550/file/ru/",
    "previousComparable": "https://stat.gov.kz/api/iblock/element/400792/file/ru/",
    "annual": "https://stat.gov.kz/api/iblock/element/472191/file/ru/",
}


# Сопоставление СЗПТ с кодами ТН ВЭД ЕАЭС из официальной таблицы БНС.
# Подход консервативный: где товар в статистике шире потребительского товара,
# в note/sourceScope это явно фиксируется для дальнейшей ручной доводки.
TRADE_CODE_RULES_2025: dict[str, dict[str, Any]] = {
    "рис шлифованный": {
        "prefixes": ["100630"],
        "scope": "полуобрушенный или полностью обрушенный рис",
    },
    "крупа гречневая": {
        "prefixes": ["110319", "110429"],
        "include_any": ["греч"],
        "fallback_prefixes": ["100810"],
        "scope": "гречневая крупа; при отсутствии отдельной строки используется гречиха",
    },
    "мука пшеничная первого сорта": {
        "prefixes": ["110100"],
        "scope": "мука пшеничная или пшенично-ржаная",
    },
    "хлеб пшеничный из муки первого сорта": {
        "prefixes": ["1905"],
        "include_any": ["хлеб"],
        "scope": "хлеб и хлебобулочные изделия; код шире розничного товара",
    },
    "рожки": {
        "prefixes": ["1902"],
        "scope": "макаронные изделия",
    },
    "говядина с костями": {
        "prefixes": ["0201", "0202"],
        "scope": "говядина свежая, охлажденная и мороженая; групповой код",
    },
    "говядина бескостная": {
        "prefixes": ["0201", "0202"],
        "scope": "говядина свежая, охлажденная и мороженая; групповой код",
    },
    "конина, включая бескостную": {
        "prefixes": ["020500"],
        "scope": "мясо лошадей, ослов, мулов или лошаков",
    },
    "баранина, включая бескостную": {
        "prefixes": ["0204"],
        "scope": "баранина или козлятина; групповой код",
    },
    "мясо птицы": {
        "prefixes": ["0207"],
        "scope": "мясо и пищевые субпродукты домашней птицы; групповой код",
    },
    "мясной фарш": {
        "prefixes": ["0201", "0202"],
        "scope": "говядина как ближайшая торговая группа; требуется уточнение кода фарша",
    },
    "рыба свежая, охлажденная, мороженая (лещ, карась, судак, карп, сазан)": {
        "prefixes": ["0302", "0303"],
        "scope": "рыба свежая, охлажденная и мороженая",
    },
    "молоко (пастеризованное, ультрапастеризованное, стерилизованное от 2,2% до 6% жирности)": {
        "prefixes": ["040120"],
        "scope": "молоко и сливки с жирностью более 1%, но не более 6%",
    },
    "сыр твердый, полутвердый": {
        "prefixes": ["040690"],
        "scope": "прочие сыры; ближайшая группа для твердых/полутвердых сыров",
    },
    "творог 5-9% жирности": {
        "exact": ["0406105001", "0406105002"],
        "scope": "творог, выделенный в 10-значных кодах",
    },
    "сметана": {
        "prefixes": ["040390"],
        "scope": "ферментированные/сквашенные молоко и сливки; групповой код, не только сметана",
    },
    "кефир 2-3% жирности": {
        "prefixes": ["040320"],
        "scope": "йогурт, пахта, кефир и прочие ферментированные продукты; групповой код",
    },
    "яйца, 1 категории": {
        "prefixes": ["040721"],
        "quantity": "supplement",
        "unit": "тыс. штук",
        "priceUnit": "$/тыс. шт.",
        "scope": "яйца кур домашних свежие; объем берется из дополнительной единицы измерения",
    },
    "масло подсолнечное": {
        "prefixes": ["1512"],
        "scope": "масло подсолнечное, сафлоровое или хлопковое и их фракции",
    },
    "масло сливочное": {
        "prefixes": ["0405"],
        "scope": "сливочное масло и прочие жиры из молока",
    },
    "яблоки": {
        "prefixes": ["080810"],
        "scope": "яблоки свежие",
    },
    "капуста белокочанная": {
        "exact": ["0704901001"],
        "scope": "капуста белокочанная",
    },
    "огурцы": {
        "prefixes": ["0707"],
        "scope": "огурцы и корнишоны свежие или охлажденные",
    },
    "помидоры": {
        "prefixes": ["0702"],
        "scope": "томаты свежие или охлажденные",
    },
    "лук репчатый": {
        "prefixes": ["070310"],
        "scope": "лук репчатый и шалот",
    },
    "морковь": {
        "prefixes": ["070610"],
        "scope": "морковь и репа",
    },
    "картофель": {
        "prefixes": ["0701"],
        "scope": "картофель свежий или охлажденный",
    },
    "сахар-песок": {
        "prefixes": ["170199"],
        "scope": "прочий сахар в твердом состоянии, ТН ВЭД 170199",
    },
    "соль, кроме экстра": {
        "prefixes": ["2501"],
        "scope": "соль и хлорид натрия",
    },
    "чай черный": {
        "prefixes": ["0902"],
        "scope": "чай",
    },
}


def norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().casefold()


PRICE_UNITS = {
    norm(name): "тг/кг"
    for name in SZPT_ORDER
}
PRICE_UNITS[norm("яйца, 1 категории")] = "тг/10 шт."

STOCK_PRODUCT_TO_SZPT = {
    "Рис": ["рис шлифованный"],
    "Крупа гречневая": ["крупа гречневая"],
    "Мука пшеничная": ["мука пшеничная первого сорта"],
    "Хлеб": ["хлеб пшеничный из муки первого сорта"],
    "Макароны": ["рожки"],
    "Говядина": ["говядина с костями"],
    "Конина": ["конина, включая бескостную"],
    "Баранина": ["баранина, включая бескостную"],
    "Куры": ["мясо птицы"],
    "Мясо кур": ["мясо птицы"],
    "Рыба": ["рыба свежая, охлажденная, мороженая (лещ, карась, судак, карп, сазан)"],
    "Молоко обработанное": ["молоко (пастеризованное, ультрапастеризованное, стерилизованное от 2,2% до 6% жирности)"],
    "Сыры и творог": ["сыр твердый, полутвердый", "творог 5-9% жирности"],
    "Кефир": ["кефир 2-3% жирности"],
    "Яйцо куриное": ["яйца, 1 категории"],
    "Масло подсолнечное": ["масло подсолнечное"],
    "Масло сливочное": ["масло сливочное"],
    "Яблоки": ["яблоки"],
    "Капуста": ["капуста белокочанная"],
    "Огурцы": ["огурцы"],
    "Томаты": ["помидоры"],
    "Лук репчатый": ["лук репчатый"],
    "Морковь столовая": ["морковь"],
    "Картофель": ["картофель"],
    "Сахар-песок": ["сахар-песок"],
    "Соль": ["соль, кроме экстра"],
}

STOCK_CATEGORY_LABELS = {
    "agricultural_enterprises": "сельхозпредприятия",
    "farms": "КФХ",
    "other_enterprises": "торговые точки и ИП",
    "warehouses": "склады/ОРЦ/ТЛЦ",
    "stabilization": "стабфонды",
    "vegetable_storage": "овощехранилища",
    "fruit_storage": "фруктохранилища",
}


def legacy_price_key(value: object) -> str:
    text = product_key(value)
    text = re.sub(r"\s*\d+\)\s*$", "", text)
    text = re.sub(r"\s*\((?:весовой|весовая)\)\s*", " ", text)
    text = re.sub(r",\s*(?:литр|десяток)$", "", text)
    return norm(text)


def relaxed_price_key(value: object) -> str:
    text = legacy_price_key(value).replace("ё", "е")
    text = re.sub(r"\([^)]*\)", " ", text)
    for phrase in [
        "с костями",
        "бескостная",
        "бескостную",
        "включая",
        "свежие",
        "несоленое",
        "ядрица",
    ]:
        text = text.replace(phrase, " ")
    return norm(text)


def resolve_price_regions(snapshot: Any, name: str) -> tuple[dict[str, float] | None, str | None]:
    alias = PRICE_ALIASES.get(name, name)
    candidates = []
    for value in [alias, name]:
        for fn in [product_key, legacy_price_key, relaxed_price_key]:
            candidate = fn(value)
            if candidate and candidate not in candidates:
                candidates.append(candidate)
    for key in candidates:
        if key in snapshot.prices:
            return snapshot.prices[key], key
    relaxed_candidates = {relaxed_price_key(value) for value in [alias, name]}
    relaxed_candidates = {value for value in relaxed_candidates if len(value) >= 5}
    best_key = None
    best_score = -1
    for key in snapshot.prices:
        relaxed_key = relaxed_price_key(key)
        if not relaxed_key:
            continue
        score = -1
        for candidate in relaxed_candidates:
            if candidate == relaxed_key:
                score = max(score, 1000 + len(candidate))
            elif len(candidate) >= 8 and (candidate in relaxed_key or relaxed_key in candidate):
                score = max(score, 500 + min(len(candidate), len(relaxed_key)))
        if score > best_score:
            best_key = key
            best_score = score
    if best_key is not None and best_score >= 500:
        return snapshot.prices[best_key], best_key
    return None, None


def safe_float(value: object) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)) and math.isfinite(value):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace("\u00a0", "").replace(" ", "").replace(",", ".").strip()
        if not cleaned:
            return None
        try:
            number = float(cleaned)
        except ValueError:
            return None
        return number if math.isfinite(number) else None
    return None


def clean_hs_code(value: object) -> str:
    return re.sub(r"\D", "", str(value or ""))


def trade_price_per_ton(value_usd_thousand: float | None, tons: float | None) -> float | None:
    if value_usd_thousand is None or tons in (None, 0):
        return None
    return round(value_usd_thousand * 1000 / tons, 2)


def trade_quantity(row: dict[str, Any], flow: str, rule: dict[str, Any]) -> float:
    if rule.get("quantity") == "supplement":
        value = safe_float(row.get(f"{flow}Supplement"))
        return value or 0.0
    value = safe_float(row.get(f"{flow}Tons"))
    return value or 0.0


def trade_unit(rule: dict[str, Any]) -> str:
    return str(rule.get("unit") or "тонн")


def trade_price_unit(rule: dict[str, Any]) -> str:
    return str(rule.get("priceUnit") or "$/т")


def trade_code_matches(code: str, name: str, rule: dict[str, Any], *, use_fallback: bool = False) -> bool:
    exact = set(rule.get("exact", []))
    prefixes = list(rule.get("fallback_prefixes" if use_fallback else "prefixes", []))
    if exact and code not in exact:
        return False
    if prefixes and not any(code.startswith(prefix) for prefix in prefixes):
        return False
    if not exact and not prefixes:
        return False
    lowered = name.casefold()
    include_any = [str(value).casefold() for value in rule.get("include_any", [])]
    include_all = [str(value).casefold() for value in rule.get("include_all", [])]
    exclude_any = [str(value).casefold() for value in rule.get("exclude_any", [])]
    if include_any and not any(value in lowered for value in include_any):
        return False
    if include_all and not all(value in lowered for value in include_all):
        return False
    if exclude_any and any(value in lowered for value in exclude_any):
        return False
    return True


def summarize_trade_rows(rows: list[dict[str, Any]], rule: dict[str, Any]) -> dict[str, Any] | None:
    if not rows:
        return None
    export_qty = sum(trade_quantity(row, "export", rule) for row in rows)
    import_qty = sum(trade_quantity(row, "import", rule) for row in rows)
    export_usd = sum(row["exportUsdK"] for row in rows)
    import_usd = sum(row["importUsdK"] for row in rows)
    codes = sorted({row["code"] for row in rows})
    names = []
    for row in rows:
        if row["name"] not in names:
            names.append(row["name"])
        if len(names) >= 6:
            break
    return {
        "export2025": round(export_qty, 5),
        "import2025": round(import_qty, 5),
        "exportUsdK2025": round(export_usd, 5),
        "importUsdK2025": round(import_usd, 5),
        "exportUsdPerTon2025": trade_price_per_ton(export_usd, export_qty),
        "importUsdPerTon2025": trade_price_per_ton(import_usd, import_qty),
        "tradeSource": OFFICIAL_TRADE_SOURCE_2025,
        "tradeSourceUrl": OFFICIAL_TRADE_URL_2025,
        "tradePeriod": "январь–декабрь 2025",
        "tradeUnit": trade_unit(rule),
        "tradePriceUnit": trade_price_unit(rule),
        "tradeCodes": codes,
        "tradeScope": rule.get("scope"),
        "tradeMatchedRows": len(rows),
        "tradeMatchedNames": names,
    }


def summarize_current_trade_rows(
    rows: list[dict[str, Any]],
    rule: dict[str, Any],
    *,
    period: str,
    source: str,
    source_url: str,
) -> dict[str, Any] | None:
    if not rows:
        return None
    month_order: list[str] = []
    monthly: dict[str, dict[str, float]] = {}
    for row in rows:
        for point in row.get("monthly", []):
            month = str(point.get("month") or "").strip()
            if not month:
                continue
            if month not in monthly:
                monthly[month] = {
                    "export": 0.0,
                    "import": 0.0,
                    "exportUsdK": 0.0,
                    "importUsdK": 0.0,
                }
                month_order.append(month)
            monthly[month]["export"] += trade_quantity(point, "export", rule)
            monthly[month]["import"] += trade_quantity(point, "import", rule)
            monthly[month]["exportUsdK"] += safe_float(point.get("exportUsdK")) or 0.0
            monthly[month]["importUsdK"] += safe_float(point.get("importUsdK")) or 0.0
    if not month_order:
        month_order = [period]
        monthly[period] = {
            "export": sum(trade_quantity(row, "export", rule) for row in rows),
            "import": sum(trade_quantity(row, "import", rule) for row in rows),
            "exportUsdK": sum(safe_float(row.get("exportUsdK")) or 0.0 for row in rows),
            "importUsdK": sum(safe_float(row.get("importUsdK")) or 0.0 for row in rows),
        }
    codes = sorted({row["code"] for row in rows})
    names = []
    for row in rows:
        if row["name"] not in names:
            names.append(row["name"])
        if len(names) >= 6:
            break
    return {
        "trade2026": {
            "months": month_order,
            "export": [round(monthly[month]["export"], 5) for month in month_order],
            "import": [round(monthly[month]["import"], 5) for month in month_order],
            "exportUsdPerTon": [
                trade_price_per_ton(monthly[month]["exportUsdK"], monthly[month]["export"])
                for month in month_order
            ],
            "importUsdPerTon": [
                trade_price_per_ton(monthly[month]["importUsdK"], monthly[month]["import"])
                for month in month_order
            ],
            "exportUsdK": [round(monthly[month]["exportUsdK"], 5) for month in month_order],
            "importUsdK": [round(monthly[month]["importUsdK"], 5) for month in month_order],
            "source": source,
            "sourceUrl": source_url,
            "codes": codes,
            "scope": rule.get("scope"),
            "unit": trade_unit(rule),
            "priceUnit": trade_price_unit(rule),
            "matchedRows": len(rows),
            "matchedNames": names,
        },
        "tradeCurrentPeriod": period,
        "tradeCurrentSource": source,
        "tradeCurrentSourceUrl": source_url,
    }


def trade_month_label(value: object) -> str:
    text = re.sub(r"\s+", " ", str(value or "").replace("_", " ")).strip()
    text = text.replace("*", "")
    text = re.sub(r"\s*г\.?\s*$", "", text, flags=re.IGNORECASE).strip()
    parts = text.split()
    return parts[0] if parts else text


def current_trade_metric_columns(sheet) -> list[dict[str, Any]]:
    groups: list[dict[str, Any]] = []
    current_group: dict[str, Any] | None = None
    current_flow = ""
    for col in range(1, sheet.max_column + 1):
        month_cell = sheet.cell(4, col).value
        if month_cell:
            label = trade_month_label(month_cell)
            current_group = {
                "month": label,
                "exportTons": [],
                "exportSupplement": [],
                "exportUsdK": [],
                "importTons": [],
                "importSupplement": [],
                "importUsdK": [],
            }
            groups.append(current_group)
            current_flow = ""
        flow_cell = sheet.cell(5, col).value
        if flow_cell:
            current_flow = norm(flow_cell)
        unit = norm(sheet.cell(6, col).value)
        if current_group is None or not current_flow or not unit:
            continue
        is_export = "экспорт" in current_flow
        is_import = "импорт" in current_flow
        is_tons = "тонн" in unit
        is_supplement = "доп" in unit
        is_usd = "доллар" in unit
        if is_export and is_tons:
            current_group["exportTons"].append(col - 1)
        elif is_export and is_supplement:
            current_group["exportSupplement"].append(col - 1)
        elif is_export and is_usd:
            current_group["exportUsdK"].append(col - 1)
        elif is_import and is_tons:
            current_group["importTons"].append(col - 1)
        elif is_import and is_supplement:
            current_group["importSupplement"].append(col - 1)
        elif is_import and is_usd:
            current_group["importUsdK"].append(col - 1)
    return [
        group for group in groups
        if group["exportTons"] or group["exportSupplement"] or group["exportUsdK"]
        or group["importTons"] or group["importSupplement"] or group["importUsdK"]
    ]


def sum_trade_columns(row: tuple[Any, ...], indexes: list[int]) -> float:
    total = 0.0
    for index in indexes:
        value = safe_float(row[index] if index < len(row) else None)
        if value is not None:
            total += value
    return total


def iter_current_trade_rows(workbook) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for sheet_name in workbook.sheetnames:
        if "ТН ВЭД" not in sheet_name:
            continue
        sheet = workbook[sheet_name]
        month_groups = current_trade_metric_columns(sheet)
        if not month_groups:
            continue
        for row in sheet.iter_rows(min_row=7, values_only=True):
            code = clean_hs_code(row[0] if len(row) > 0 else None)
            if len(code) < 4:
                continue
            name = str(row[1] if len(row) > 1 and row[1] is not None else "").strip()
            monthly = []
            for group in month_groups:
                point = {
                    "month": group["month"],
                    "exportTons": sum_trade_columns(row, group["exportTons"]),
                    "exportSupplement": sum_trade_columns(row, group["exportSupplement"]),
                    "exportUsdK": sum_trade_columns(row, group["exportUsdK"]),
                    "importTons": sum_trade_columns(row, group["importTons"]),
                    "importSupplement": sum_trade_columns(row, group["importSupplement"]),
                    "importUsdK": sum_trade_columns(row, group["importUsdK"]),
                }
                if any(
                    point[key] != 0
                    for key in ("exportTons", "exportSupplement", "exportUsdK", "importTons", "importSupplement", "importUsdK")
                ):
                    monthly.append(point)
            if not monthly:
                continue
            export_tons = sum(point["exportTons"] for point in monthly)
            export_supplement = sum(point["exportSupplement"] for point in monthly)
            export_usd = sum(point["exportUsdK"] for point in monthly)
            import_tons = sum(point["importTons"] for point in monthly)
            import_supplement = sum(point["importSupplement"] for point in monthly)
            import_usd = sum(point["importUsdK"] for point in monthly)
            rows.append(
                {
                    "code": code,
                    "name": name,
                    "exportTons": export_tons,
                    "exportSupplement": export_supplement,
                    "exportUsdK": export_usd,
                    "importTons": import_tons,
                    "importSupplement": import_supplement,
                    "importUsdK": import_usd,
                    "monthly": monthly,
                }
            )
    return rows


def current_trade_rule_can_use(code: str, rule: dict[str, Any], *, use_fallback: bool = False) -> bool:
    exact = [str(value) for value in rule.get("exact", [])]
    prefixes = [str(value) for value in rule.get("fallback_prefixes" if use_fallback else "prefixes", [])]
    if exact:
        return code in exact
    if not prefixes:
        return False
    max_prefix_len = max(len(prefix) for prefix in prefixes)
    needs_text_filter = any(rule.get(key) for key in ("include_any", "include_all", "exclude_any"))
    if needs_text_filter:
        return len(code) >= max_prefix_len
    return len(code) == max_prefix_len


def load_official_trade_2025() -> dict[str, dict[str, Any]]:
    if not OFFICIAL_TRADE_FILE_2025.exists():
        print(f"Официальная торговая выгрузка БНС не найдена: {OFFICIAL_TRADE_FILE_2025}")
        return {}
    workbook = load_workbook(OFFICIAL_TRADE_FILE_2025, data_only=True, read_only=True)
    try:
        rows_by_product: dict[str, list[dict[str, Any]]] = {product_id(name): [] for name in TRADE_CODE_RULES_2025}
        fallback_by_product: dict[str, list[dict[str, Any]]] = {product_id(name): [] for name in TRADE_CODE_RULES_2025}
        exact_rules: dict[str, list[tuple[str, dict[str, Any]]]] = {}
        prefix_rules: list[tuple[str, str, dict[str, Any], bool]] = []
        for product_name, rule in TRADE_CODE_RULES_2025.items():
            pid = product_id(product_name)
            for code in rule.get("exact", []):
                exact_rules.setdefault(code, []).append((pid, rule))
            for prefix in rule.get("prefixes", []):
                prefix_rules.append((prefix, pid, rule, False))
            for prefix in rule.get("fallback_prefixes", []):
                prefix_rules.append((prefix, pid, rule, True))
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, values_only=True):
                code = clean_hs_code(row[0] if len(row) > 0 else None)
                if len(code) < 4:
                    continue
                name = str(row[1] if len(row) > 1 and row[1] is not None else "").strip()
                export_tons = safe_float(row[3] if len(row) > 3 else None) or 0.0
                export_supplement = safe_float(row[4] if len(row) > 4 else None) or 0.0
                export_usd = safe_float(row[5] if len(row) > 5 else None) or 0.0
                import_tons = safe_float(row[6] if len(row) > 6 else None) or 0.0
                import_supplement = safe_float(row[7] if len(row) > 7 else None) or 0.0
                import_usd = safe_float(row[8] if len(row) > 8 else None) or 0.0
                if (
                    export_tons == 0 and export_supplement == 0 and export_usd == 0
                    and import_tons == 0 and import_supplement == 0 and import_usd == 0
                ):
                    continue
                trade_row = {
                    "code": code,
                    "name": name,
                    "exportTons": export_tons,
                    "exportSupplement": export_supplement,
                    "exportUsdK": export_usd,
                    "importTons": import_tons,
                    "importSupplement": import_supplement,
                    "importUsdK": import_usd,
                }
                candidates: list[tuple[str, dict[str, Any], bool]] = []
                candidates.extend((pid, rule, False) for pid, rule in exact_rules.get(code, []))
                candidates.extend(
                    (pid, rule, is_fallback)
                    for prefix, pid, rule, is_fallback in prefix_rules
                    if code.startswith(prefix)
                )
                seen: set[tuple[str, bool]] = set()
                for pid, rule, is_fallback in candidates:
                    key = (pid, is_fallback)
                    if key in seen:
                        continue
                    seen.add(key)
                    if trade_code_matches(code, name, rule):
                        rows_by_product[pid].append(trade_row)
                    elif is_fallback and trade_code_matches(code, name, rule, use_fallback=True):
                        fallback_by_product[pid].append(trade_row)
        result: dict[str, dict[str, Any]] = {}
        for product_name, rule in TRADE_CODE_RULES_2025.items():
            pid = product_id(product_name)
            rows = rows_by_product.get(pid) or fallback_by_product.get(pid) or []
            summary = summarize_trade_rows(rows, rule)
            if summary:
                if not rows_by_product.get(pid) and fallback_by_product.get(pid):
                    summary["tradeScope"] = f"{summary.get('tradeScope')}; использован fallback-код"
                result[pid] = summary
        return result
    finally:
        workbook.close()


def load_official_trade_current() -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    if not OFFICIAL_TRADE_CURRENT_META.exists():
        return {}, {}
    try:
        meta = json.loads(OFFICIAL_TRADE_CURRENT_META.read_text(encoding="utf-8"))
    except Exception as exc:
        print(f"Не удалось прочитать метаданные текущей торговли БНС: {exc}")
        return {}, {}
    file_name = str(meta.get("file") or "")
    trade_file = OFFICIAL_TRADE_CURRENT_DIR / file_name
    if not file_name or not trade_file.exists():
        print(f"Текущий файл торговли БНС не найден: {trade_file}")
        return {}, meta
    period = str(meta.get("period") or meta.get("title") or "2026")
    source = str(meta.get("title") or "БНС: Экспорт и импорт товаров РК по 4,6,10 знакам ТН ВЭД ЕАЭС")
    source_url = str(meta.get("url") or OFFICIAL_TRADE_URL_CURRENT)
    workbook = load_workbook(trade_file, data_only=True, read_only=True)
    try:
        rows_by_product: dict[str, list[dict[str, Any]]] = {product_id(name): [] for name in TRADE_CODE_RULES_2025}
        fallback_by_product: dict[str, list[dict[str, Any]]] = {product_id(name): [] for name in TRADE_CODE_RULES_2025}
        exact_rules: dict[str, list[tuple[str, dict[str, Any]]]] = {}
        prefix_rules: list[tuple[str, str, dict[str, Any], bool]] = []
        for product_name, rule in TRADE_CODE_RULES_2025.items():
            pid = product_id(product_name)
            for code in rule.get("exact", []):
                exact_rules.setdefault(code, []).append((pid, rule))
            for prefix in rule.get("prefixes", []):
                prefix_rules.append((prefix, pid, rule, False))
            for prefix in rule.get("fallback_prefixes", []):
                prefix_rules.append((prefix, pid, rule, True))
        for trade_row in iter_current_trade_rows(workbook):
            code = trade_row["code"]
            name = trade_row["name"]
            candidates: list[tuple[str, dict[str, Any], bool]] = []
            candidates.extend((pid, rule, False) for pid, rule in exact_rules.get(code, []))
            candidates.extend(
                (pid, rule, is_fallback)
                for prefix, pid, rule, is_fallback in prefix_rules
                if code.startswith(prefix)
            )
            seen: set[tuple[str, bool]] = set()
            for pid, rule, is_fallback in candidates:
                key = (pid, is_fallback)
                if key in seen:
                    continue
                seen.add(key)
                if (
                    current_trade_rule_can_use(code, rule)
                    and trade_code_matches(code, name, rule)
                ):
                    rows_by_product[pid].append(trade_row)
                elif (
                    is_fallback
                    and current_trade_rule_can_use(code, rule, use_fallback=True)
                    and trade_code_matches(code, name, rule, use_fallback=True)
                ):
                    fallback_by_product[pid].append(trade_row)
        result: dict[str, dict[str, Any]] = {}
        for product_name, rule in TRADE_CODE_RULES_2025.items():
            pid = product_id(product_name)
            rows = rows_by_product.get(pid) or fallback_by_product.get(pid) or []
            summary = summarize_current_trade_rows(
                rows,
                rule,
                period=period,
                source=source,
                source_url=source_url,
            )
            if summary:
                if not rows_by_product.get(pid) and fallback_by_product.get(pid):
                    summary["trade2026"]["scope"] = f"{summary['trade2026'].get('scope')}; использован fallback-код"
                result[pid] = summary
        return result, meta
    finally:
        workbook.close()


def apply_balance_from_official_sources(item: dict[str, Any], trade: dict[str, Any] | None) -> None:
    item["export2025"] = None
    item["import2025"] = None
    item["exportUsdK2025"] = None
    item["importUsdK2025"] = None
    item["exportUsdPerTon2025"] = None
    item["importUsdPerTon2025"] = None
    item["coverage"] = None
    item["importShare"] = None
    item["consumption"] = None
    item["coverageSourceProduct"] = None
    item["coverageFormula"] = "производство / (производство − экспорт + импорт) × 100"
    item["coverageSource"] = "расчет по официальным данным БНС"
    item["balanceUnit"] = trade.get("tradeUnit") if trade else item.get("balanceUnit")
    if not trade:
        item["tradeSource"] = None
        item["tradeScope"] = "нет уверенного сопоставления с кодом ТН ВЭД"
        return
    item.update(trade)
    item["balanceUnit"] = trade.get("tradeUnit") or item.get("balanceUnit") or "тонн"
    production = safe_float(item.get("balanceProduction"))
    export = safe_float(trade.get("export2025"))
    import_value = safe_float(trade.get("import2025"))
    if production is None or export is None or import_value is None:
        return
    consumption = production - export + import_value
    item["consumption"] = round(consumption, 5)
    if consumption > 0:
        item["coverage"] = round(production / consumption * 100, 10)
        item["importShare"] = round(import_value / consumption * 100, 10)


def empty_macro() -> dict[str, Any]:
    return {
        "period": "н/д",
        "inflation": None,
        "monthlyInflation": None,
        "foodInflation": None,
        "foodContribution": None,
        "monthlyFoodInflation": None,
        "publicationDate": None,
        "nextPublicationDate": None,
        "sourceTitle": None,
        "sourceUrl": None,
        "inflationFoot": "",
        "foodFoot": "",
        "contributionFoot": "",
        "monthlyFoodFoot": "",
        "comparison2025": {
            "inflation": None,
            "foodInflation": None,
            "foodContribution": None,
        },
        "groups": [],
    }


def official_macro() -> dict[str, Any]:
    # БНС → Экономика → Статистика цен → Публикации:
    # "Инфляция в Республике Казахстан (июнь 2026г.)", опубликовано 01.07.2026.
    # Источник фиксируется здесь, чтобы верхний блок не показывал временные заглушки.
    return {
        "period": "июнь 2026",
        "publicationDate": "2026-07-01",
        "nextPublicationDate": "2026-08-03",
        "sourceTitle": "БНС: Инфляция в Республике Казахстан (июнь 2026г.)",
        "sourceUrl": "https://stat.gov.kz/ru/industries/economy/prices/publications/347710/",
        "inflation": 10.3,
        "monthlyInflation": 0.8,
        "foodInflation": 10.4,
        "foodContribution": 4.354,
        "monthlyFoodInflation": 0.6,
        "inflationFoot": "за месяц: +0,8%; в мае 2026: 10,4%",
        "foodFoot": "за месяц: +0,6%; в мае 2026: 10,7%",
        "contributionFoot": "продовольственные товары, вклад в годовую инфляцию",
        "monthlyFoodFoot": "месячный рост продовольственных товаров",
        "comparison2025": {
            "inflation": 11.8,
            "foodInflation": None,
            "foodContribution": None,
        },
        "groups": [
            {"name": "Продовольственные товары", "inflation": 10.4, "contribution": 4.354},
            {"name": "Мясо и птица", "inflation": 17.6, "contribution": 1.555},
            {"name": "Колбасы, изделия из мяса", "inflation": 16.3, "contribution": 0.542},
            {"name": "Хлеб и хлебобулочные изделия", "inflation": 8.8, "contribution": 0.281},
            {"name": "Безалкогольные напитки", "inflation": 13.8, "contribution": 0.250},
            {"name": "Масла и жиры", "inflation": 10.7, "contribution": 0.236},
            {"name": "Сыр и творог", "inflation": 11.0, "contribution": 0.189},
            {"name": "Яйца", "inflation": 15.3, "contribution": 0.183},
            {"name": "Кисломолочные продукты", "inflation": 13.3, "contribution": 0.180},
            {"name": "Молоко питьевое", "inflation": 10.9, "contribution": 0.145},
            {"name": "Фрукты, ягоды и орехи", "inflation": 1.5, "contribution": 0.078},
            {"name": "Сахар", "inflation": 3.5, "contribution": 0.020},
            {"name": "Крупы", "inflation": 5.0, "contribution": 0.001},
            {"name": "Овощи", "inflation": -11.9, "contribution": -0.277},
        ],
    }


def load_base() -> dict[str, Any]:
    # Strict source mode: do not reuse the previous dashboard JSON as a data source.
    # It may contain stale local/manual values from earlier iterations.
    return {
        "meta": {
            "updated": date.today().isoformat(),
            "productCount": 0,
            "sourcePolicy": "official_bns_taldau_only",
        },
        "macro": official_macro(),
        "products": [],
    }


def load_operational_stocks() -> dict[str, Any]:
    empty = {
        "enabled": False,
        "source": "Telegram-бот / оперативные данные МИО",
        "sourceNote": "Не является официальной статистикой БНС; используется как оперативный слой запасов от регионов.",
        "updated": None,
        "expectedRegions": len(STOCK_REGIONS),
        "receivedRegions": 0,
        "missingRegions": list(STOCK_REGIONS),
        "latestReports": [],
        "productRows": [],
        "byProductId": {},
        "totalTons": None,
        "eggThousandPieces": None,
    }
    if not STOCKS_DB.exists():
        return empty

    connection = sqlite3.connect(STOCKS_DB)
    connection.row_factory = sqlite3.Row
    try:
        latest_reports = connection.execute(
            """SELECT r.* FROM reports r JOIN (
                 SELECT region, MAX(report_date) report_date FROM reports GROUP BY region
               ) latest ON latest.region=r.region AND latest.report_date=r.report_date
               ORDER BY r.region"""
        ).fetchall()
        if not latest_reports:
            return empty

        latest_report_ids = [row["id"] for row in latest_reports]
        placeholders = ",".join("?" for _ in latest_report_ids)
        stock_rows = connection.execute(
            f"""
            SELECT r.region, r.report_date, s.product, s.unit, s.total,
                   s.agricultural_enterprises, s.farms, s.other_enterprises, s.warehouses,
                   s.stabilization_direct, s.stabilization_forward, s.stabilization_revolving,
                   s.vegetable_storage, s.fruit_storage
            FROM reports r
            JOIN stock_rows s ON s.report_id = r.id
            WHERE r.id IN ({placeholders})
            ORDER BY s.product, r.region
            """,
            latest_report_ids,
        ).fetchall()
    finally:
        connection.close()

    received = {row["region"] for row in latest_reports}
    report_dates = [row["report_date"] for row in latest_reports if row["report_date"]]
    updated = max(report_dates) if report_dates else None
    latest_reports_payload = [
        {
            "region": row["region"],
            "reportDate": row["report_date"],
            "sourceFile": row["source_file"],
            "receivedAt": row["received_at"],
        }
        for row in latest_reports
    ]

    grouped: dict[str, dict[str, Any]] = {}
    for row in stock_rows:
        product = row["product"]
        unit = "тыс. шт." if product == "Яйцо куриное" else "тонн"
        item = grouped.setdefault(
            product,
            {
                "product": product,
                "unit": unit,
                "total": 0.0,
                "latestDate": row["report_date"],
                "regionCount": 0,
                "regions": [],
                "categoryTotals": {key: 0.0 for key in STOCK_CATEGORY_LABELS},
                "matchedProductIds": [product_id(name) for name in STOCK_PRODUCT_TO_SZPT.get(product, [])],
                "sourceScope": "оперативная группа МИО",
            },
        )
        value = safe_float(row["total"]) or 0.0
        item["total"] += value
        item["regionCount"] += 1
        item["latestDate"] = max(item["latestDate"], row["report_date"])
        item["regions"].append(
            {
                "region": row["region"],
                "reportDate": row["report_date"],
                "total": round(value, 3),
            }
        )
        item["categoryTotals"]["agricultural_enterprises"] += safe_float(row["agricultural_enterprises"]) or 0.0
        item["categoryTotals"]["farms"] += safe_float(row["farms"]) or 0.0
        item["categoryTotals"]["other_enterprises"] += safe_float(row["other_enterprises"]) or 0.0
        item["categoryTotals"]["warehouses"] += safe_float(row["warehouses"]) or 0.0
        item["categoryTotals"]["stabilization"] += (
            (safe_float(row["stabilization_direct"]) or 0.0)
            + (safe_float(row["stabilization_forward"]) or 0.0)
            + (safe_float(row["stabilization_revolving"]) or 0.0)
        )
        item["categoryTotals"]["vegetable_storage"] += safe_float(row["vegetable_storage"]) or 0.0
        item["categoryTotals"]["fruit_storage"] += safe_float(row["fruit_storage"]) or 0.0

    product_rows = []
    by_product_id: dict[str, dict[str, Any]] = {}
    total_tons = 0.0
    egg_thousand = 0.0
    for item in grouped.values():
        item["total"] = round(item["total"], 3)
        item["regions"].sort(key=lambda row: row["total"], reverse=True)
        item["categoryTotals"] = {
            key: round(value, 3)
            for key, value in item["categoryTotals"].items()
            if abs(value) >= 0.0005
        }
        if item["unit"] == "тыс. шт.":
            egg_thousand += item["total"]
        else:
            total_tons += item["total"]
        product_rows.append(item)
        for pid in item["matchedProductIds"]:
            by_product_id[pid] = item

    product_rows.sort(key=lambda row: row["total"], reverse=True)
    return {
        "enabled": True,
        "source": "Telegram-бот / оперативные данные МИО",
        "sourceNote": "Не является официальной статистикой БНС; используется как оперативный слой запасов от регионов.",
        "updated": updated,
        "expectedRegions": len(STOCK_REGIONS),
        "receivedRegions": len(received),
        "missingRegions": [region for region in STOCK_REGIONS if region not in received],
        "latestReports": latest_reports_payload,
        "productRows": product_rows,
        "byProductId": by_product_id,
        "totalTons": round(total_tons, 3) if total_tons else None,
        "eggThousandPieces": round(egg_thousand, 3) if egg_thousand else None,
    }


def product_id(name: str) -> str:
    return norm(name)


def latest_file(patterns: list[str], min_size: int = 1) -> Path | None:
    candidates: list[Path] = []
    search_roots = [
        RAW_DIR,
        RAW_DIR / "converted",
        RAW_DIR / "official",
        RAW_DIR / "industrial_production",
    ]
    for pattern in patterns:
        for root in search_roots:
            if root.exists():
                candidates.extend(root.glob(pattern))
    candidates = [
        p for p in candidates
        if p.is_file() and not p.name.startswith("~$") and p.suffix.lower() == ".xlsx" and p.stat().st_size >= min_size
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda p: (p.stat().st_mtime, p.stat().st_size))


def sheet3(path: Path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["3 "] if "3 " in workbook.sheetnames else workbook[workbook.sheetnames[2]]
    return workbook, sheet


def parse_production_table(path: Path, kind: str) -> dict[str, dict[str, Any]]:
    workbook, sheet = sheet3(path)
    try:
        rows: dict[str, dict[str, Any]] = {}
        for row in range(1, sheet.max_row + 1):
            name = str(sheet.cell(row, 1).value or "").strip()
            if not name:
                continue
            if kind == "monthly":
                rows[norm(name)] = {
                    "name": name,
                    "previousMonth": safe_float(sheet.cell(row, 2).value),
                    "monthValue": safe_float(sheet.cell(row, 3).value),
                    "currentPeriod": safe_float(sheet.cell(row, 4).value),
                    "previousYearPeriod": safe_float(sheet.cell(row, 6).value),
                    "changePct": safe_float(sheet.cell(row, 9).value),
                }
            else:
                rows[norm(name)] = {
                    "name": name,
                    "year2025": safe_float(sheet.cell(row, 2).value),
                    "year2024": safe_float(sheet.cell(row, 3).value),
                    "changePct": safe_float(sheet.cell(row, 4).value),
                }
        return rows
    finally:
        workbook.close()


def find_rows(rows: dict[str, dict[str, Any]], patterns: list[str]) -> list[dict[str, Any]]:
    result = []
    used = set()
    for pattern in patterns:
        needle = norm(pattern)
        match_key = next((key for key in rows if needle in key), None)
        if match_key and match_key not in used:
            result.append(rows[match_key])
            used.add(match_key)
    return result


def sum_field(rows: list[dict[str, Any]], field: str) -> float | None:
    values = [row.get(field) for row in rows if row.get(field) is not None]
    if not values:
        return None
    return float(sum(values))


def pct_ratio(value: float | None, previous: float | None) -> float | None:
    if value is None or previous in (None, 0):
        return None
    return round(value / previous * 100, 1)


def read_taldau_cache(product_id_value: str) -> dict[str, Any] | None:
    if not OFFICIAL_CACHE_ALLOWED:
        return None
    if not TALDAU_CACHE.exists():
        return None
    try:
        data = json.loads(TALDAU_CACHE.read_text(encoding="utf-8"))
        item = data.get(product_id_value)
        if isinstance(item, dict) and str(item.get("source", "")).startswith("Taldau"):
            item["cacheStatus"] = "cached"
            return item
    except Exception:
        return None
    return None


def write_taldau_cache(product_id_value: str, value: dict[str, Any]) -> None:
    if not OFFICIAL_CACHE_ALLOWED:
        return
    try:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        data = json.loads(TALDAU_CACHE.read_text(encoding="utf-8")) if TALDAU_CACHE.exists() else {}
        data[product_id_value] = value
        TALDAU_CACHE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as exc:
        print(f"Taldau cache not written for {product_id_value}: {exc}")


def latest_converted_file(name_parts: list[str]) -> Path | None:
    if not (RAW_DIR / "converted").exists():
        return None
    candidates = []
    for path in (RAW_DIR / "converted").glob("*.xlsx"):
        key = norm(path.name)
        if all(norm(part) in key for part in name_parts):
            candidates.append(path)
    return max(candidates, key=lambda p: (p.stat().st_mtime, p.stat().st_size)) if candidates else None


def discover_price_source_files(directory: Path) -> list[Any]:
    if not directory.exists():
        return []
    sources = []
    for path in sorted(directory.iterdir()):
        if not path.is_file() or path.suffix.lower() != ".xlsx":
            continue
        name = norm(path.name)
        if "bns_prices_" not in name and "15-05" not in name:
            continue
        workbook = load_workbook(path, data_only=True, read_only=True)
        try:
            period = _observation_date(workbook, path)
        finally:
            workbook.close()
        sources.append(SimpleNamespace(observation_date=period, path=path))
    return sources


def load_detailed_vegetable_harvest() -> dict[str, Any]:
    official_dir = RAW_DIR / "crop_harvest"
    year_files: dict[int, Path] = {}
    for year in [2024, 2025]:
        official = official_dir / f"T-03-15-G-2tom-{year}.xlsx"
        if official.exists():
            year_files[year] = official
    if 2025 not in year_files:
        return {}

    mapping = {
        "картофель": [["картофел"]],
        "морковь": [["морков"]],
        "лук репчатый": [["лука репчат"]],
        "капуста белокочанная": [["капуст"]],
        "помидоры": [["помидор"]],
        "огурцы": [["огурц"]],
    }

    def parse_year(path: Path) -> dict[str, dict[str, Any]]:
        result: dict[str, dict[str, Any]] = {}
        workbook = load_workbook(path, data_only=True, read_only=True)
        try:
            if "7" not in workbook.sheetnames:
                return result
            sheet = workbook["7"]
            sections: list[dict[str, Any]] = []
            for row in range(1, sheet.max_row + 1):
                title = sheet.cell(row, 1).value
                if not isinstance(title, str) or not title.strip().startswith("7."):
                    continue
                rk_row = None
                for lookup_row in range(row + 1, min(row + 14, sheet.max_row + 1)):
                    label = norm(sheet.cell(lookup_row, 1).value)
                    if "республика казахстан" in label:
                        rk_row = lookup_row
                        break
                value_centners = safe_float(sheet.cell(rk_row, 2).value) if rk_row else None
                if value_centners is None:
                    continue
                sections.append(
                    {
                        "title": re.sub(r"\s+", " ", title).strip(),
                        "norm": norm(title),
                        "row": row,
                        "valueTons": value_centners / 10,
                    }
                )

            def section_sum(terms: list[str]) -> tuple[float | None, list[str]]:
                matched = [
                    section for section in sections
                    if all(term in section["norm"] for term in terms)
                ]
                if not matched:
                    return None, []
                total = sum(section["valueTons"] for section in matched)
                rows = [f"{section['title']} (строка {section['row']})" for section in matched]
                return total, rows

            for product_name, term_groups in mapping.items():
                total = 0.0
                rows: list[str] = []
                for terms in term_groups:
                    value, source_rows = section_sum(terms)
                    if value is None:
                        continue
                    total += value
                    rows.extend(source_rows)
                if rows:
                    result[product_name] = {
                        "value": total,
                        "rows": rows,
                        "source": "БНС: Т-03-15-Г (2-том), валовой сбор овощей",
                    }

            if "11" in workbook.sheetnames:
                fruit_sheet = workbook["11"]
                for row in range(1, fruit_sheet.max_row + 1):
                    title = fruit_sheet.cell(row, 1).value
                    if not isinstance(title, str) or "валовый сбор яблок" not in norm(title):
                        continue
                    rk_row = None
                    for lookup_row in range(row + 1, min(row + 16, fruit_sheet.max_row + 1)):
                        label = norm(fruit_sheet.cell(lookup_row, 1).value)
                        if "республика казахстан" in label:
                            rk_row = lookup_row
                            break
                    value_centners = safe_float(fruit_sheet.cell(rk_row, 2).value) if rk_row else None
                    if value_centners is None:
                        continue
                    result["яблоки"] = {
                        "value": value_centners / 10,
                        "rows": [f"{re.sub(r'\\s+', ' ', title).strip()} (строка {row})"],
                        "source": "БНС: Т-03-15-Г (2-том), валовой сбор плодов",
                    }
                    break
            return result
        finally:
            workbook.close()

    parsed = {year: parse_year(path) for year, path in year_files.items()}
    output: dict[str, Any] = {}
    for product_name, current in parsed.get(2025, {}).items():
        previous = parsed.get(2024, {}).get(product_name)
        current_value = current["value"]
        previous_value = previous["value"] if previous else None
        output[product_id(product_name)] = {
            "source": current["source"],
            "sourceFiles": {
                "current": year_files[2025].name,
                "annual": year_files[2025].name,
                "previous": year_files.get(2024).name if 2024 in year_files else None,
            },
            "sourceUrls": {
                "current": OFFICIAL_CROP_HARVEST_URLS.get(2025),
                "annual": OFFICIAL_CROP_HARVEST_URLS.get(2025),
                "previous": OFFICIAL_CROP_HARVEST_URLS.get(2024) if 2024 in year_files else None,
            },
            "sourceRows": current["rows"],
            "unit": "тонн",
            "current": {
                "period": "2025 год",
                "value": current_value,
                "previousYearPeriod": previous_value,
                "changePct": pct_ratio(current_value, previous_value),
                "monthValue": None,
                "previousMonth": None,
            },
            "annual": {
                "2025": current_value,
                "2024": previous_value,
                "changePct": pct_ratio(current_value, previous_value),
            },
        }
    return output


def load_crop_harvest() -> dict[str, Any]:
    return load_detailed_vegetable_harvest()


def load_bns_production() -> tuple[dict[str, Any], dict[str, str]]:
    monthly_file = latest_file(["T-04-01-М_рус*.xlsx", "Т-04-01-М_рус*.xlsx"], min_size=200_000)
    annual_file = latest_file(["Т-04-05-Г_рус*.xlsx", "T-04-05-Г_рус*.xlsx"], min_size=200_000)
    monthly = parse_production_table(monthly_file, "monthly") if monthly_file else {}
    annual = parse_production_table(annual_file, "annual") if annual_file else {}
    output: dict[str, Any] = {}
    for name in SZPT_ORDER:
        patterns = PRODUCTION_ROWS.get(name)
        if not patterns:
            continue
        monthly_rows = find_rows(monthly, patterns)
        annual_rows = find_rows(annual, patterns)
        current_value = sum_field(monthly_rows, "currentPeriod")
        previous_value = sum_field(monthly_rows, "previousYearPeriod")
        year2025 = sum_field(annual_rows, "year2025")
        year2024 = sum_field(annual_rows, "year2024")
        if current_value is None and year2025 is None:
            continue
        output[product_id(name)] = {
            "source": "БНС: Т-04-01-М и Т-04-05-Г",
            "sourceFiles": {
                "current": monthly_file.name if monthly_file else None,
                "annual": annual_file.name if annual_file else None,
            },
            "sourceRows": sorted({row["name"] for row in monthly_rows + annual_rows}),
            "unit": "тонн",
            "current": {
                "period": "январь–май 2026",
                "value": current_value,
                "previousYearPeriod": previous_value,
                "changePct": pct_ratio(current_value, previous_value),
                "monthValue": sum_field(monthly_rows, "monthValue"),
                "previousMonth": sum_field(monthly_rows, "previousMonth"),
            },
            "annual": {
                "2025": year2025,
                "2024": year2024,
                "changePct": pct_ratio(year2025, year2024),
            },
        }
    sugar_pid = product_id("сахар-песок")
    buckwheat_pid = product_id("крупа гречневая")
    taldau_sugar = fetch_taldau_sugar_production()
    if taldau_sugar:
        write_taldau_cache(sugar_pid, taldau_sugar)
    else:
        taldau_sugar = read_taldau_cache(sugar_pid)
    if taldau_sugar:
        output[sugar_pid] = taldau_sugar
    taldau_buckwheat = fetch_taldau_accumulated_production(
        keyword="крупа гречневая",
        term_id="18749915",
        source_row="РЕСПУБЛИКА КАЗАХСТАН + Всего + Крупа и мука грубого помола гречневая",
    )
    if taldau_buckwheat:
        write_taldau_cache(buckwheat_pid, taldau_buckwheat)
    else:
        taldau_buckwheat = read_taldau_cache(buckwheat_pid)
    if taldau_buckwheat:
        output[buckwheat_pid] = taldau_buckwheat
    livestock = load_livestock_production()
    output.update(livestock)
    crop_harvest = load_crop_harvest()
    output.update(crop_harvest)
    return output, {
        "monthlyProductionFile": monthly_file.name if monthly_file else "",
        "annualProductionFile": annual_file.name if annual_file else "",
        "sugarProductionSource": "Taldau 701608" if taldau_sugar else "",
        "buckwheatProductionSource": "Taldau 701608" if taldau_buckwheat else "",
        "livestockProductionSource": "БНС Т-03-02-М" if livestock else "",
        "cropHarvestSource": "БНС валовой сбор" if crop_harvest else "",
    }


def fetch_taldau_sugar_production() -> dict[str, Any] | None:
    """Load sugar production directly from Taldau index 701608.

    Taldau segment:
    - indexId 701608: produced industrial goods in physical terms
    - periodId 8: month with accumulation
    - terms: Kazakhstan + total by enterprise size + sugar product
    """
    return fetch_taldau_accumulated_production(
        keyword="сахар-песок",
        term_id="4122296",
        source_row="РЕСПУБЛИКА КАЗАХСТАН + Всего + Сахар-песок, полученный из сахара тростникового или свекловичного, в твердом состоянии",
    )


def fetch_taldau_accumulated_production(keyword: str, term_id: str, source_row: str) -> dict[str, Any] | None:
    url = "https://taldau.stat.gov.kz/ru/NewIndex/GetIndexTreeData"
    params = {
        "p_index_id": "701608",
        "p_keyword": keyword,
        "p_period_id": "8",
        "p_measure_id": "0",
        "p_parent_id": "",
        "p_term_id": "741880",
        "p_terms": f"741880,741927,{term_id}",
        "p_dicIds": "68,90,1973",
        "idx": "2",
    }
    rows = None
    last_error = None
    for _ in range(2):
        try:
            request = urllib.request.Request(
                url,
                data=urllib.parse.urlencode(params).encode("utf-8"),
                method="POST",
                headers={
                    "User-Agent": "Mozilla/5.0",
                    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                },
            )
            with urllib.request.urlopen(request, timeout=35) as response:
                rows = json.loads(response.read().decode("utf-8"))
            break
        except Exception as exc:
            last_error = exc
    if rows is None:
        print(f"Taldau {keyword} недоступен, официальные данные пропущены: {last_error}")
        return None
    if not rows:
        return None
    row = rows[0]
    keys = sorted(
        (
            key for key in row
            if re.fullmatch(r"y\d{6}", key) and safe_float(row.get(key)) is not None
        ),
        key=lambda key: (int(key[3:7]), int(key[1:3])),
    )
    if not keys:
        return None
    latest = keys[-1]
    month = latest[1:3]
    year = latest[3:7]
    previous_same = f"y{month}{int(year) - 1}"
    current_value = safe_float(row.get(latest))
    previous_value = safe_float(row.get(previous_same))
    year2025 = safe_float(row.get("y122025"))
    year2024 = safe_float(row.get("y122024"))
    month_names = {
        "01": "январь",
        "02": "январь–февраль",
        "03": "январь–март",
        "04": "январь–апрель",
        "05": "январь–май",
        "06": "январь–июнь",
        "07": "январь–июль",
        "08": "январь–август",
        "09": "январь–сентябрь",
        "10": "январь–октябрь",
        "11": "январь–ноябрь",
        "12": "январь–декабрь",
    }
    return {
        "source": "Taldau.stat.gov.kz: индекс 701608",
        "sourceFiles": {
            "current": f"https://taldau.stat.gov.kz/ru/NewIndex/GetIndex/701608?keyword={urllib.parse.quote(keyword)}",
            "annual": f"https://taldau.stat.gov.kz/ru/NewIndex/GetIndex/701608?keyword={urllib.parse.quote(keyword)}",
        },
        "sourceRows": [source_row],
        "unit": "тонн",
        "current": {
            "period": f"{month_names.get(month, month)} {year}",
            "value": current_value,
            "previousYearPeriod": previous_value,
            "changePct": pct_ratio(current_value, previous_value),
            "monthValue": None,
            "previousMonth": None,
        },
        "annual": {
            "2025": year2025,
            "2024": year2024,
            "changePct": pct_ratio(year2025, year2024),
        },
    }


def latest_official_livestock_file(patterns: list[str]) -> Path | None:
    official_dir = RAW_DIR / "livestock"
    candidates: list[Path] = []
    if official_dir.exists():
        for pattern in patterns:
            candidates.extend(official_dir.glob(pattern))
    candidates = [p for p in candidates if p.is_file() and not p.name.startswith("~$")]
    return max(candidates, key=lambda p: p.stat().st_mtime) if candidates else None


def rk_row(sheet) -> int | None:
    for row in range(1, min(sheet.max_row, 30) + 1):
        value = norm(sheet.cell(row, 1).value)
        if "республика казахстан" in value:
            return row
    return None


def read_livestock_file(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        meat = workbook["2.4"]
        eggs = workbook["4"]
        meat_row = rk_row(meat)
        egg_row = rk_row(eggs)
        if meat_row is None or egg_row is None:
            return {}
        return {
            "beef": safe_float(meat.cell(meat_row, 3).value),
            "lamb": safe_float(meat.cell(meat_row, 4).value),
            "horse": safe_float(meat.cell(meat_row, 7).value),
            "poultry": safe_float(meat.cell(meat_row, 9).value),
            "eggs": safe_float(eggs.cell(egg_row, 14).value),
            "eggsPrevious": safe_float(eggs.cell(egg_row, 15).value),
            "eggUnit": "тыс. штук",
        }
    finally:
        workbook.close()


def load_livestock_production() -> dict[str, Any]:
    current_file = latest_official_livestock_file(["Т-03-02-М 05 2026 рус*.xlsx", "T-03-02-М 05 2026 рус*.xlsx"])
    previous_file = latest_official_livestock_file(["Т-03-02-М 05 2025 рус*.xlsx", "T-03-02-М 05 2025 рус*.xlsx"])
    annual_file = latest_official_livestock_file(["Т-03-02-М 12 2025 рус*.xlsx", "T-03-02-М 12 2025 рус*.xlsx"])
    if not current_file:
        return {}
    current = read_livestock_file(current_file)
    previous = read_livestock_file(previous_file) if previous_file else {}
    annual = read_livestock_file(annual_file) if annual_file else {}
    if not current:
        return {}
    mapping = {
        "говядина с костями": ("beef", "крупный рогатый скот, убойный вес"),
        "говядина бескостная": ("beef", "крупный рогатый скот, убойный вес"),
        "конина, включая бескостную": ("horse", "лошади, убойный вес"),
        "баранина, включая бескостную": ("lamb", "овцы, убойный вес"),
        "мясо птицы": ("poultry", "птица, убойный вес"),
        "яйца, 1 категории": ("eggs", "получено яиц куриных, все категории хозяйств"),
    }
    result: dict[str, Any] = {}
    for product_name, (field, row_label) in mapping.items():
        value = current.get(field)
        prev = previous.get(field)
        annual_2025 = annual.get(field)
        annual_2024 = annual.get(f"{field}Previous")
        if value is None:
            continue
        unit = "тыс. штук" if field == "eggs" else "тонн"
        result[product_id(product_name)] = {
            "source": "БНС: Т-03-02-М",
            "sourceFiles": {
                "current": current_file.name,
                "annual": annual_file.name if annual_file else None,
                "previousComparable": previous_file.name if previous_file else None,
            },
            "sourceUrls": {
                "current": OFFICIAL_LIVESTOCK_URLS["current"],
                "annual": OFFICIAL_LIVESTOCK_URLS["annual"] if annual_file else None,
                "previousComparable": OFFICIAL_LIVESTOCK_URLS["previousComparable"] if previous_file else None,
                "page": OFFICIAL_LIVESTOCK_PAGE_URL,
            },
            "sourceRows": [row_label],
            "unit": unit,
            "balanceCompatible": True,
            "current": {
                "period": "январь–май 2026",
                "value": value,
                "previousYearPeriod": prev,
                "changePct": pct_ratio(value, prev),
                "monthValue": None,
                "previousMonth": None,
            },
            "annual": {
                "2025": annual_2025,
                "2024": annual_2024,
                "changePct": pct_ratio(annual_2025, annual_2024),
            },
        }
    return result


def load_price_history() -> tuple[dict[str, list[dict[str, Any]]], dict[str, Any]]:
    files = []
    for directory in [RAW_DIR, *(p for p in RAW_DIR.iterdir() if p.is_dir())] if RAW_DIR.exists() else []:
        try:
            files.extend(discover_price_source_files(directory))
        except Exception as exc:
            print(f"Пропущена папка цен {directory}: {exc}")
    unique = {}
    for source in files:
        existing = unique.get(source.observation_date)
        if existing is None or source.path.stat().st_mtime > existing.path.stat().st_mtime:
            unique[source.observation_date] = source
    files = sorted(unique.values(), key=lambda item: item.observation_date, reverse=True)
    snapshots = []
    for source in files:
        try:
            snapshots.append(parse_snapshot(source))
        except Exception as exc:
            try:
                snapshots.append(parse_legacy_price_snapshot(source))
            except Exception as legacy_exc:
                print(f"Пропущен файл цен {source.path.name}: {exc}; legacy: {legacy_exc}")
    snapshots.sort(key=lambda item: item.source.observation_date)
    history: dict[str, list[dict[str, Any]]] = {product_id(name): [] for name in SZPT_ORDER}
    price_by_year: dict[str, dict[str, list[dict[str, Any]]]] = {
        product_id(name): {"2024": [], "2025": [], "2026": []} for name in SZPT_ORDER
    }
    latest_by_product: dict[str, dict[str, Any]] = {}
    for snapshot in snapshots:
        iso = snapshot.source.observation_date.isoformat()
        for name in SZPT_ORDER:
            pid = product_id(name)
            regions, key = resolve_price_regions(snapshot, name)
            if not regions:
                continue
            price = regions.get("По обследованным городам") or regions.get("Республика Казахстан")
            if price is None:
                price = sum(regions.values()) / len(regions)
            city_regions = {
                region: value for region, value in regions.items()
                if region not in {"По обследованным городам", "Республика Казахстан"}
            }
            point_date = snapshot.source.observation_date
            point = {"date": iso, "price": round(float(price), 1)}
            history[pid].append(point)
            year = str(point_date.year)
            if year in price_by_year[pid]:
                price_by_year[pid][year].append({
                    "date": iso,
                    "month": point_date.month,
                    "dayOfYear": point_date.timetuple().tm_yday,
                    "label": point_date.strftime("%d.%m"),
                    "price": round(float(price), 1),
                })
            latest_by_product[pid] = {
                "price": round(float(price), 1),
                "regions": city_regions,
                "weekChange": snapshot.week_change.get(key) if key else None,
                "yearChange": snapshot.year_change.get(key) if key else None,
                "annualChange": snapshot.annual_change.get(key) if key else None,
            }
    meta = {
        "priceFiles": [s.source.path.name for s in snapshots],
        "latestPriceDate": snapshots[-1].source.observation_date.isoformat() if snapshots else date.today().isoformat(),
        "latestByProduct": latest_by_product,
        "priceByYear": price_by_year,
    }
    return history, meta


def parse_legacy_price_snapshot(source) -> Any:
    workbook = load_workbook(source.path, data_only=True, read_only=True)
    try:
        if "5" not in workbook.sheetnames:
            raise ValueError("нет листа 5 со средними ценами")
        price_sheet = workbook["5"]
        index_sheet = workbook["1"] if "1" in workbook.sheetnames else None
        header_row = 3
        regions = {
            col: str(price_sheet.cell(header_row, col).value).strip()
            for col in range(2, price_sheet.max_column + 1)
            if price_sheet.cell(header_row, col).value
        }
        prices: dict[str, dict[str, float]] = {}
        for row in range(header_row + 1, price_sheet.max_row + 1):
            name = price_sheet.cell(row, 1).value
            if not name:
                continue
            values = {}
            for col, region in regions.items():
                value = safe_float(price_sheet.cell(row, col).value)
                if value is not None:
                    values[region.strip()] = value
            if values:
                prices[legacy_price_key(name)] = values
        annual_change: dict[str, float] = {}
        year_change: dict[str, float] = {}
        week_change: dict[str, float] = {}
        if index_sheet is not None:
            for row in range(5, index_sheet.max_row + 1):
                name = index_sheet.cell(row, 1).value
                if not name:
                    continue
                key = legacy_price_key(name)
                annual = safe_float(index_sheet.cell(row, 2).value)
                year = safe_float(index_sheet.cell(row, 3).value)
                week = safe_float(index_sheet.cell(row, 5).value)
                if annual is not None:
                    annual_change[key] = round(annual - 100, 1)
                if year is not None:
                    year_change[key] = round(year - 100, 1)
                if week is not None:
                    week_change[key] = round(week - 100, 1)
        if not prices:
            raise ValueError("старый шаблон распознан, но цены не найдены")
        return SimpleNamespace(
            source=source,
            prices=prices,
            annual_change=annual_change,
            year_change=year_change,
            week_change=week_change,
        )
    finally:
        workbook.close()


def risk_label(score: float) -> str:
    if score >= 55:
        return "Высокий риск"
    if score >= 28:
        return "Средний риск"
    return "Низкий риск"


def recalc_risk(item: dict[str, Any]) -> None:
    coverage = item.get("coverage")
    import_share = item.get("importShare") or 0
    annual = item.get("annualChange") or 0
    week = item.get("weekChange") or 0
    score = 0.0
    if coverage is None:
        score += 18
    elif coverage < 50:
        score += 34
    elif coverage < 80:
        score += 22
    elif coverage < 100:
        score += 12
    score += min(max(import_share, 0), 120) * 0.18
    score += max(annual, 0) * 0.75
    score += max(week, 0) * 2.2
    item["riskScore"] = round(min(score, 100), 1)
    item["riskLabel"] = risk_label(item["riskScore"])


def main() -> None:
    base = load_base()
    operational_stocks = load_operational_stocks()
    history, price_meta = load_price_history()
    production, production_meta = load_bns_production()
    official_trade = load_official_trade_2025()
    current_trade, current_trade_meta = load_official_trade_current()
    products = []
    for number, name in enumerate(SZPT_ORDER, start=1):
        pid = product_id(name)
        item: dict[str, Any] = {
            "id": pid,
            "coverage": None,
            "importShare": None,
            "consumption": None,
            "export2025": None,
            "import2025": None,
            "exportUsd2025": None,
            "importUsd2025": None,
            "exportUsdPerTon2025": None,
            "importUsdPerTon2025": None,
        }
        item["number"] = number
        item["name"] = name[:1].upper() + name[1:]
        item["isSzpt"] = True
        price_info = price_meta["latestByProduct"].get(pid, {})
        item["history"] = history.get(pid, [])
        item["priceByYear"] = price_meta.get("priceByYear", {}).get(pid, {})
        item["price"] = price_info.get("price")
        item["regions"] = price_info.get("regions", {})
        item["weekChange"] = price_info.get("weekChange")
        item["yearChange"] = price_info.get("yearChange")
        item["annualChange"] = price_info.get("annualChange")
        item["unit"] = PRICE_UNITS.get(pid, "тг/кг")
        item["operationalStocksMio"] = operational_stocks.get("byProductId", {}).get(pid)
        prod = production.get(pid)
        item["productionBns"] = prod
        if prod:
            current_value = prod.get("current", {}).get("value")
            annual_2025 = prod.get("annual", {}).get("2025")
            balance_compatible = prod.get("balanceCompatible", prod.get("unit") == "тонн")
            display_production = annual_2025 if annual_2025 is not None else None
            balance_production = annual_2025 if annual_2025 is not None and balance_compatible else None
            item["balanceProduction"] = balance_production
            item["production"] = display_production
            item["productionPeriod"] = "2025 год" if display_production is not None else None
            item["productionBalanceCompatible"] = bool(balance_compatible)
            item["operationalProduction"] = current_value
            item["operationalProductionPeriod"] = prod.get("current", {}).get("period")
            item["seasonalProduction"] = None
        else:
            item["production"] = None
            item["balanceProduction"] = None
            item["productionPeriod"] = None
            item["productionBalanceCompatible"] = False
            item["operationalProduction"] = None
            item["operationalProductionPeriod"] = None
            item["seasonalProduction"] = None
        apply_balance_from_official_sources(item, official_trade.get(pid))
        if pid in current_trade:
            item.update(current_trade[pid])
        item.setdefault("trade2026", {"months": [], "export": [], "import": [], "exportUsdPerTon": [], "importUsdPerTon": []})
        recalc_risk(item)
        products.append(item)
    base["products"] = sorted(products, key=lambda x: x["riskScore"], reverse=True)
    stocks_public = dict(operational_stocks)
    stocks_public.pop("byProductId", None)
    base["stocks"] = stocks_public
    base["meta"].update(
        {
            "updated": price_meta["latestPriceDate"],
            "productCount": len(products),
            "sourcePolicy": "official_bns_taldau_plus_mio_operational_stocks",
            "sourcePolicyNote": "Официальные показатели собираются из БНС/Талдау; оперативные запасы МИО показываются отдельным слоем из Telegram-бота и не смешиваются с официальной статистикой.",
            "priceSeriesMode": "weekly_bns",
            "priceSourceNote": "График цен строится только по загруженным недельным публикациям БНС Т-15-05-Н.",
            "priceSourceUrl": BNS_PRICE_SOURCE_PAGE,
            "tradeSource2025": OFFICIAL_TRADE_SOURCE_2025 if official_trade else "",
            "tradeSourceUrl2025": OFFICIAL_TRADE_URL_2025 if official_trade else "",
            "tradeProductsMapped2025": len(official_trade),
            "tradeSource2026": current_trade_meta.get("title", "") if current_trade else "",
            "tradeSourceUrl2026": current_trade_meta.get("url", "") if current_trade else "",
            "tradePeriod2026": current_trade_meta.get("period", "") if current_trade else "",
            "tradeProductsMapped2026": len(current_trade),
            "livestockSourceUrl": OFFICIAL_LIVESTOCK_PAGE_URL if production_meta.get("livestockProductionSource") else "",
            "cropHarvestSourceUrls": OFFICIAL_CROP_HARVEST_URLS,
            "coverageFormula": "производство / (производство − экспорт + импорт) × 100",
            "weeklyUpdatePlan": "цены СЗПТ — пятница после 12:00; экспорт/импорт — ежемесячно после 15-го; производство БНС/Талдау — ежедневная проверка после 15-го числа в 15:00",
            "mioStocksSource": "Telegram-бот / data/stocks.sqlite3",
            "mioStocksUpdated": operational_stocks.get("updated"),
            **production_meta,
        }
    )
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(base, ensure_ascii=False, indent=2), encoding="utf-8")
    public = ROOT / "dashboard-cloudflare-pages"
    if public.exists():
        for source in ["index.html", "app.js", "styles.css"]:
            shutil.copy2(DASHBOARD_DIR / source, public / source)
        (public / "data").mkdir(exist_ok=True)
        shutil.copy2(OUTPUT_JSON, public / "data" / "bns.json")
    print(f"Готово: {OUTPUT_JSON}")
    print(f"Цены: {price_meta['latestPriceDate']}; производство: {production_meta}")


if __name__ == "__main__":
    main()
