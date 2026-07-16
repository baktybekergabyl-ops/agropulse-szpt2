import shutil
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from .models import Snapshot
from .parser import key, product_key

PRODUCT_ALIASES = {
    "капуста белокочанная": "Капуста белокочанная",
    "картофель": "Картофель",
    "лук репчатый": "Лук репчатый",
    "баранина": "Баранина, включая бескостную",
    "сахар-песок": "Сахар-песок",
    "морковь": "Морковь",
    "мясо кур": "Мясо кур (бедро, голень, окорочка куриные)",
    "масло сливочное": "Масло сливочное",
    "конина": "Конина, включая бескостную",
    "рожки": "Рожки",
    "говядина с костями": "Говядина с костями",
    "крупа гречневая": "Крупа гречневая",
    "хлеб": "Хлеб пшеничный из муки первого сорта",
    "говядина бескостная": "Говядина бескостная",
    "мясной фарш": "Мясной фарш",
    "сыр": "Сыр твердый, полутвердый",
    "мука": "Мука пшеничная первого сорта",
    "куры": "Куры",
    "масло подсолнечное": "Масло подсолнечное, литр",
    "рис шлифованный": "Рис шлифованный",
    "рыба": "Рыба свежая, охлажденная, мороженая (лещ, карась, судак, карп, сазан)",
    "сметана": "Сметана",
    "яблоки": "Яблоки",
    "соль": "Соль, кроме экстра",
    "молоко": "Молоко (пастеризованное, ультрапастеризованное, стерилизованное от 2,2% до 6% жирности), литр",
    "творог 5-9% жирности": "Творог 5-9% жирности",
    "кефир 2-3% жирности": "Кефир 2-3% жирности, литр",
    "чай черный": "Чай черный",
    "яйца, 1 категории": "Яйца, 1 категории, десяток",
    "огурцы": "Огурцы",
    "помидоры": "Помидоры",
}

REGION_NAMES = {
    "Астана": "Астана",
    "Алматы": "г. Алматы",
    "Шымкент": "г. Шымкент",
    "Актау": "Мангистауская",
    "Актобе": "Актюбинская",
    "Атырау": "Атырауская",
    "Жезказган": "Улытау",
    "Кокшетау": "Акмолинская",
    "Караганда": "Карагандинская",
    "Конаев": "Алматинская",
    "Костанай": "Костанайская",
    "Кызылорда": "Кызылординская",
    "Уральск": "ЗКО",
    "Усть-Каменогорск": "ВКО",
    "Павлодар": "Павлодарская",
    "Петропавловск": "СКО",
    "Семей": "Абай",
    "Талдыкорган": "Жетису",
    "Тараз": "Жамбылская",
    "Туркестан": "Туркестанская",
}


def _short_date(value: date) -> str:
    return value.strftime("%d.%m")


def _long_date(value: date) -> str:
    months = [
        "", "ЯНВАРЯ", "ФЕВРАЛЯ", "МАРТА", "АПРЕЛЯ", "МАЯ", "ИЮНЯ",
        "ИЮЛЯ", "АВГУСТА", "СЕНТЯБРЯ", "ОКТЯБРЯ", "НОЯБРЯ", "ДЕКАБРЯ",
    ]
    return f"{value.day} {months[value.month]} {value.year} ГОДА"


def _product_key(template_name: str) -> str:
    source_name = PRODUCT_ALIASES.get(key(template_name), template_name)
    return product_key(source_name)


def _national_price(snapshot: Snapshot | None, product_key: str) -> float | None:
    if snapshot is None:
        return None
    values = snapshot.prices.get(product_key, {})
    for region, value in values.items():
        if key(region).startswith("по обследованным городам"):
            return value
    return None


def _sort_products_by_week(
    product_names: list[str],
    snapshot: Snapshot,
) -> list[str]:
    def sort_key(product_name: str) -> tuple[bool, float]:
        value = snapshot.week_change.get(_product_key(product_name))
        return value is not None, value if value is not None else float("-inf")

    return sorted(product_names, key=sort_key, reverse=True)


def _extreme(values: dict[str, float], minimum: bool) -> str | None:
    regional = [
        (region, value) for region, value in values.items()
        if not key(region).startswith("по обследованным городам")
    ]
    if not regional:
        return None
    region, price = (min if minimum else max)(regional, key=lambda item: item[1])
    label = REGION_NAMES.get(region, region)
    formatted = f"{price:,.0f}".replace(",", " ")
    return f"{label} - {formatted} тг"


def create_report(
    template_path: Path,
    current: Snapshot,
    previous: Snapshot | None,
    output_dir: Path,
) -> Path:
    if not template_path.exists():
        raise FileNotFoundError(f"Основной шаблон не найден: {template_path}")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / (
        f"bns_price_analysis_{current.source.observation_date.isoformat()}.xlsx"
    )
    shutil.copy2(template_path, output_path)
    workbook = load_workbook(output_path)
    sheet = workbook[workbook.sheetnames[0]]

    current_date = current.source.observation_date
    sheet["B2"] = f"СРЕДНИЕ ЦЕНЫ НА СЗПТ ПО ДАННЫМ СТАТИСТИКИ НА {_long_date(current_date)}"
    previous_label = (
        _short_date(previous.source.observation_date) if previous else "—"
    )
    sheet["F4"] = f"Цена по РК (тг) на\n{previous_label}"
    sheet["G4"] = f"Цена по РК (тг) на\n{_short_date(current_date)}"

    product_names = [
        str(sheet.cell(row, 3).value or "")
        for row in range(5, 36)
    ]
    product_names = _sort_products_by_week(product_names, current)

    # The template highlights rows 5-9, so sorting makes them the weekly top five.
    for row, product_name in zip(range(5, 36), product_names):
        product_key = _product_key(product_name)
        values = current.prices.get(product_key, {})
        sheet.cell(row, 2).value = row - 4
        sheet.cell(row, 3).value = product_name
        sheet.cell(row, 4).value = current.year_change.get(product_key)
        sheet.cell(row, 5).value = current.week_change.get(product_key)
        sheet.cell(row, 6).value = _national_price(previous, product_key)
        sheet.cell(row, 7).value = _national_price(current, product_key)
        sheet.cell(row, 8).value = _extreme(values, minimum=True)
        sheet.cell(row, 9).value = _extreme(values, minimum=False)

    workbook.save(output_path)
    return output_path
