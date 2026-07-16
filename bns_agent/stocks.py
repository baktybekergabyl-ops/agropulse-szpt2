from __future__ import annotations

import re
import sqlite3
import csv
import io
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


REGIONS = (
    "Абай", "Акмолинская", "Актюбинская", "Алматинская", "Атырауская",
    "Восточно-Казахстанская", "Жамбылская", "Жетісу",
    "Западно-Казахстанская", "Карагандинская", "Костанайская",
    "Кызылординская", "Мангистауская", "Павлодарская",
    "Северо-Казахстанская", "Туркестанская", "Ұлытау",
    "Астана", "Алматы", "Шымкент",
)

REGION_MARKERS = {
    "Абай": ("область абай", "абайской"),
    "Акмолинская": ("акмолинск", "акмола"),
    "Актюбинская": ("актюбинск", "актобе"),
    "Алматинская": ("алматинской области", "алм. обл"),
    "Атырауская": ("атырауск", "атырау"),
    "Восточно-Казахстанская": ("восточно-казахстан",),
    "Жамбылская": ("жамбылск", "жамбыл"),
    "Жетісу": ("жетісу", "жетысу"),
    "Западно-Казахстанская": ("западно-казахстан",),
    "Карагандинская": ("карагандинск", "караганда"),
    "Костанайская": ("костанайск", "костанай"),
    "Кызылординская": ("кызылординск",),
    "Мангистауская": ("мангистауск", "мангистау"),
    "Павлодарская": ("павлодарск", "павлодар"),
    "Северо-Казахстанская": ("северо-казахстан",),
    "Туркестанская": ("туркестанск", "туркестан"),
    "Ұлытау": ("ұлытау", "улытау"),
    "Астана": ("города астан", "г. астан", "астана"),
    "Алматы": ("города алмат", "г. алмат"),
    "Шымкент": ("шымкент",),
}

MONTHS = {
    "января": 1, "февраля": 2, "марта": 3, "апреля": 4, "мая": 5,
    "июня": 6, "июля": 7, "августа": 8, "сентября": 9, "октября": 10,
    "ноября": 11, "декабря": 12,
}
KAZAKH_MONTHS = {
    "қаңтар": 1, "ақпан": 2, "наурыз": 3, "сәуір": 4, "мамыр": 5,
    "маусым": 6, "шілде": 7, "тамыз": 8, "қыркүйек": 9, "қазан": 10,
    "қараша": 11, "желтоқсан": 12,
}

CATEGORIES = (
    "agricultural_enterprises",
    "farms",
    "other_enterprises",
    "warehouses",
    "stabilization_direct",
    "stabilization_forward",
    "stabilization_revolving",
    "vegetable_storage",
    "fruit_storage",
)

CATEGORY_LABELS = {
    "agricultural_enterprises": "Сельхозпредприятия",
    "farms": "КФХ",
    "other_enterprises": "Прочие предприятия",
    "warehouses": "Склады",
    "stabilization_direct": "Стабфонд прямой закуп",
    "stabilization_forward": "Стабфонд форвард",
    "stabilization_revolving": "Стабфонд оборотная схема",
    "vegetable_storage": "Овощехранилища",
    "fruit_storage": "Фруктохранилища",
}

STOCK_SUMMARY_TEMPLATE = Path(__file__).resolve().parents[1] / "templates" / "stocks_summary_template.xlsx"

SUMMARY_ROWS = (
    (5, "Баранина", "Баранина"),
    (6, "Говядина", "Говядина"),
    (7, "Капуста", "Капуста"),
    (8, "Картофель", "Картофель"),
    (9, "Кефир", "Кефир"),
    (10, "Колбасные изделия", "Колбасные изделия"),
    (11, "Конина", "Конина"),
    (12, "Крупа гречневая", "Крупа гречневая"),
    (13, "Лук репчатый", "Лук репчатый"),
    (14, "Макароны", "Макароны"),
    (15, "Масло подсолнечное ", "Масло подсолнечное"),
    (16, "Масло сливочное", "Масло сливочное"),
    (17, "Молоко обработанное", "Молоко обработанное"),
    (18, "Морковь столовая", "Морковь столовая"),
    (19, "Мука пшеничная", "Мука пшеничная"),
    (20, "Мясо кур", "Мясо кур"),
    (21, "Огурцы", "Огурцы"),
    (22, "Перец", "Перец"),
    (23, "Рис продовольственный", "Рис"),
    (24, "Рыба ", "Рыба"),
    (25, "Сахар-песок", "Сахар-песок"),
    (26, "Свекла столовая", "Свекла столовая"),
    (27, "Свинина", "Свинина"),
    (28, "Соль", "Соль"),
    (29, "Сыры и творог", "Сыры и творог"),
    (30, "Томаты", "Томаты"),
    (31, "Хлеб", "Хлеб"),
    (32, "Яблоки", "Яблоки"),
    (33, "Яйцо куриное, тысяч штук", "Яйцо куриное"),
)

PRODUCT_PATTERNS = (
    ("Рис", ("рис продоволь", "рис", "күріш")),
    ("Картофель", ("картофел", "картоп")),
    ("Капуста", ("капуст", "қырыққабат")),
    ("Свекла столовая", ("свекл", "свёкл", "қызылша")),
    ("Лук репчатый", ("лук репчат", "пияз")),
    ("Морковь столовая", ("морков", "сәбіз")),
    ("Перец", ("перец", "бұрыш")),
    ("Томаты", ("томат", "помидор", "қызанақ")),
    ("Огурцы", ("огур", "қияр")),
    ("Яблоки", ("яблок", "алма")),
    ("Мука пшеничная", ("мука пшен", "ұн")),
    ("Хлеб", ("хлеб", "нан")),
    ("Макароны", ("макарон",)),
    ("Крупа гречневая", ("греч", "қарақұмық")),
    ("Масло подсолнечное", ("масло подсол", "күнбағыс майы")),
    ("Масло сливочное", ("масло слив", "сары май")),
    ("Молоко обработанное", ("молоко обработ", "молоко пастер", "сүт")),
    ("Кефир", ("кефир", "айран")),
    ("Сыры и творог", ("сыры и творог", "сыр и творог", "ірімшіктер мен сүзбе")),
    ("Яйцо куриное", ("яйцо кур", "яйца кур", "тауық жұмыртқасы")),
    ("Соль", ("соль", "тұз")),
    ("Сахар-песок", ("сахар", "қант")),
    ("Говядина", ("говядин", "сиыр еті")),
    ("Свинина", ("свинин", "шошқа еті")),
    ("Баранина", ("баранин", "қой еті")),
    ("Конина", ("конин", "жылқы еті")),
    ("Мясо кур", ("мясо кур", "куриное мясо", "тауық еті")),
    ("Колбасные изделия", ("колбас", "шұжық")),
    ("Рыба", ("рыба", "балық")),
)


@dataclass(frozen=True)
class StockRow:
    product: str
    unit: str
    total: float
    categories: dict[str, float]


@dataclass(frozen=True)
class StockReport:
    region: str
    report_date: date
    source_file: str
    sheet_name: str
    rows: tuple[StockRow, ...]
    warnings: tuple[str, ...]


def _text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().casefold()


def _product(value: object) -> str | None:
    text = _text(value)
    for canonical, patterns in PRODUCT_PATTERNS:
        if any(pattern in text for pattern in patterns):
            return canonical
    return None


def _extract_date(text: str) -> date | None:
    numeric = re.search(r"(\d{1,2})[./](\d{1,2})[./](20\d{2})", text)
    if numeric:
        day, month, year = map(int, numeric.groups())
        return date(year, month, day)
    words = re.search(
        r"(\d{1,2})\s+(" + "|".join(MONTHS) + r")\s+(20\d{2})",
        text.casefold(),
    )
    if words:
        return date(int(words.group(3)), MONTHS[words.group(2)], int(words.group(1)))
    kazakh = re.search(
        r"(20\d{2})\s+жылғы\s+(\d{1,2})\s+("
        + "|".join(KAZAKH_MONTHS)
        + r")",
        text.casefold(),
    )
    if kazakh:
        return date(
            int(kazakh.group(1)),
            KAZAKH_MONTHS[kazakh.group(3)],
            int(kazakh.group(2)),
        )
    return None


def _detect_region(text: str) -> str | None:
    lowered = text.casefold()
    for region, markers in REGION_MARKERS.items():
        if any(marker in lowered for marker in markers):
            return region
    abbreviations = {
        "Восточно-Казахстанская": "вко",
        "Западно-Казахстанская": "зко",
        "Кызылординская": "кзо",
        "Северо-Казахстанская": "ско",
    }
    for region, abbreviation in abbreviations.items():
        if re.search(rf"(?<![а-яёәіңғүұқөһ]){abbreviation}(?![а-яёәіңғүұқөһ])", lowered):
            return region
    return None


def parse_stock_report(path: Path) -> StockReport:
    if path.suffix.lower() == ".xls":
        raise ValueError("Старый формат XLS не поддерживается. Сохраните файл как XLSX.")
    if path.suffix.lower() != ".xlsx":
        raise ValueError("Нужен файл Excel в формате XLSX.")

    raw = load_workbook(path, data_only=False, read_only=True)
    values = load_workbook(path, data_only=True, read_only=True)
    candidates: list[StockReport] = []
    try:
        for raw_sheet in raw.worksheets:
            value_sheet = values[raw_sheet.title]
            search_parts = [path.stem, raw_sheet.title]
            for row in raw_sheet.iter_rows(
                min_row=1, max_row=min(raw_sheet.max_row, 12),
                min_col=1, max_col=min(raw_sheet.max_column, 14),
                values_only=True,
            ):
                search_parts.extend(str(cell) for cell in row if cell is not None)
            search_text = " ".join(search_parts)
            region = _detect_region(search_text)
            report_date = _extract_date(search_text)
            if not region or not report_date:
                continue

            rows: list[StockRow] = []
            warnings: list[str] = []
            seen: set[str] = set()
            for row_number in range(1, raw_sheet.max_row + 1):
                matched = None
                product_col = None
                for col in range(1, min(raw_sheet.max_column, 14) + 1):
                    matched = _product(raw_sheet.cell(row_number, col).value)
                    if matched:
                        product_col = col
                        break
                if not matched or product_col is None or matched in seen:
                    continue
                component_start = product_col + 2
                components = []
                for col in range(component_start, component_start + len(CATEGORIES)):
                    value = value_sheet.cell(row_number, col).value
                    components.append(float(value) if isinstance(value, (int, float)) else 0.0)
                total = round(sum(components), 6)
                if total < 0:
                    warnings.append(f"{matched}: отрицательный итог")
                unit = "тыс. шт." if matched == "Яйцо куриное" else "тонн"
                rows.append(
                    StockRow(
                        product=matched,
                        unit=unit,
                        total=total,
                        categories=dict(zip(CATEGORIES, components)),
                    )
                )
                seen.add(matched)
            if len(rows) >= 20:
                missing = [name for name, _ in PRODUCT_PATTERNS if name not in seen]
                if missing:
                    warnings.append("Не найдены товары: " + ", ".join(missing))
                candidates.append(
                    StockReport(
                        region=region,
                        report_date=report_date,
                        source_file=path.name,
                        sheet_name=raw_sheet.title,
                        rows=tuple(rows),
                        warnings=tuple(warnings),
                    )
                )
    finally:
        raw.close()
        values.close()
    if not candidates:
        raise ValueError("Не удалось определить область, дату или таблицу с товарами.")
    return max(candidates, key=lambda item: item.report_date)


class StockStore:
    def __init__(self, path: Path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.connection = sqlite3.connect(path)
        self.connection.row_factory = sqlite3.Row
        self.connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS reports (
                id INTEGER PRIMARY KEY,
                region TEXT NOT NULL,
                report_date TEXT NOT NULL,
                source_file TEXT NOT NULL,
                sheet_name TEXT NOT NULL,
                received_at TEXT NOT NULL,
                telegram_chat_id INTEGER,
                warnings TEXT NOT NULL DEFAULT '',
                UNIQUE(region, report_date)
            );
            CREATE TABLE IF NOT EXISTS stock_rows (
                report_id INTEGER NOT NULL REFERENCES reports(id) ON DELETE CASCADE,
                product TEXT NOT NULL,
                unit TEXT NOT NULL,
                total REAL NOT NULL,
                agricultural_enterprises REAL NOT NULL,
                farms REAL NOT NULL,
                other_enterprises REAL NOT NULL,
                warehouses REAL NOT NULL,
                stabilization_direct REAL NOT NULL,
                stabilization_forward REAL NOT NULL,
                stabilization_revolving REAL NOT NULL,
                vegetable_storage REAL NOT NULL,
                fruit_storage REAL NOT NULL,
                PRIMARY KEY(report_id, product)
            );
            CREATE TABLE IF NOT EXISTS contacts (
                chat_id INTEGER PRIMARY KEY,
                region TEXT,
                is_admin INTEGER NOT NULL DEFAULT 0,
                active INTEGER NOT NULL DEFAULT 1
            );
            """
        )

    def save(self, report: StockReport, chat_id: int | None = None) -> None:
        with self.connection:
            old = self.connection.execute(
                "SELECT id FROM reports WHERE region=? AND report_date=?",
                (report.region, report.report_date.isoformat()),
            ).fetchone()
            if old:
                self.connection.execute("DELETE FROM stock_rows WHERE report_id=?", (old["id"],))
                report_id = old["id"]
                self.connection.execute(
                    """UPDATE reports SET source_file=?, sheet_name=?, received_at=?,
                       telegram_chat_id=?, warnings=? WHERE id=?""",
                    (
                        report.source_file, report.sheet_name, datetime.now().isoformat(timespec="seconds"),
                        chat_id, "\n".join(report.warnings), report_id,
                    ),
                )
            else:
                cursor = self.connection.execute(
                    """INSERT INTO reports
                       (region, report_date, source_file, sheet_name, received_at,
                        telegram_chat_id, warnings) VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (
                        report.region, report.report_date.isoformat(), report.source_file,
                        report.sheet_name, datetime.now().isoformat(timespec="seconds"),
                        chat_id, "\n".join(report.warnings),
                    ),
                )
                report_id = cursor.lastrowid
            for row in report.rows:
                self.connection.execute(
                    """INSERT INTO stock_rows VALUES
                       (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        report_id, row.product, row.unit, row.total,
                        *(row.categories[name] for name in CATEGORIES),
                    ),
                )

    def register_contact(self, chat_id: int, region: str | None, admin: bool = False) -> None:
        with self.connection:
            self.connection.execute(
                """INSERT INTO contacts(chat_id, region, is_admin) VALUES (?, ?, ?)
                   ON CONFLICT(chat_id) DO UPDATE SET region=excluded.region,
                   is_admin=MAX(contacts.is_admin, excluded.is_admin), active=1""",
                (chat_id, region, int(admin)),
            )

    def latest_by_region(self) -> dict[str, sqlite3.Row]:
        return {row["region"]: row for row in self.latest_reports()}

    def latest_reports(self) -> list[sqlite3.Row]:
        return self.connection.execute(
            """SELECT r.* FROM reports r JOIN (
                 SELECT region, MAX(report_date) report_date FROM reports GROUP BY region
               ) latest ON latest.region=r.region AND latest.report_date=r.report_date
               ORDER BY r.region"""
        ).fetchall()

    def contacts(self) -> list[sqlite3.Row]:
        return self.connection.execute(
            "SELECT * FROM contacts WHERE active=1 ORDER BY region"
        ).fetchall()

    def export_latest_sources_archive(self, path: Path, uploads_dir: Path) -> Path:
        path.parent.mkdir(parents=True, exist_ok=True)
        uploads_dir.mkdir(parents=True, exist_ok=True)
        used_names: set[str] = set()
        manifest: list[list[str]] = []

        def safe_name(value: object) -> str:
            cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", str(value or "").strip())
            return cleaned.strip(" .") or "unknown"

        def unique_archive_name(region: str, report_date: str, source_name: str) -> str:
            folder = safe_name(region)
            base = safe_name(source_name)
            candidate = f"{folder}/{safe_name(report_date)}_{base}"
            if candidate not in used_names:
                used_names.add(candidate)
                return candidate
            stem = Path(base).stem
            suffix = Path(base).suffix
            counter = 2
            while True:
                candidate = f"{folder}/{safe_name(report_date)}_{stem}_{counter}{suffix}"
                if candidate not in used_names:
                    used_names.add(candidate)
                    return candidate
                counter += 1

        def find_uploaded_file(source_name: str, chat_id: object) -> Path | None:
            candidates = [
                item
                for item in uploads_dir.iterdir()
                if item.is_file() and item.name.endswith(source_name)
            ]
            if not candidates:
                return None
            chat_marker = f"_{chat_id}_"

            def score(item: Path) -> tuple[int, float]:
                preferred = 1 if chat_id and chat_marker in item.name else 0
                return preferred, item.stat().st_mtime

            return max(candidates, key=score)

        with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for row in self.latest_reports():
                source_name = Path(row["source_file"]).name
                found = find_uploaded_file(source_name, row["telegram_chat_id"])
                archived_as = ""
                status = "missing"
                if found:
                    archived_as = unique_archive_name(row["region"], row["report_date"], source_name)
                    archive.write(found, archived_as)
                    status = "added"
                manifest.append(
                    [
                        row["region"],
                        row["report_date"],
                        source_name,
                        archived_as,
                        status,
                    ]
                )

            buffer = io.StringIO()
            writer = csv.writer(buffer, delimiter=";")
            writer.writerow(["region", "report_date", "source_file", "archived_as", "status"])
            writer.writerows(manifest)
            archive.writestr("manifest.csv", "\ufeff" + buffer.getvalue())
        return path

    def export_latest_summary(self, path: Path) -> Path:
        path.parent.mkdir(parents=True, exist_ok=True)
        if STOCK_SUMMARY_TEMPLATE.exists():
            workbook = load_workbook(STOCK_SUMMARY_TEMPLATE)
            sheet = workbook["Лист1"] if "Лист1" in workbook.sheetnames else workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Лист1"
            sheet.merge_cells("B2:J2")
            sheet.merge_cells("A3:J3")
            sheet.merge_cells("I35:J35")
            sheet["A3"] = "в тоннах"
            sheet["B4"] = "  Наименование товара"
            sheet["C4"] = "  Всего"
            sheet["D4"] = "  В сельскохозяйственных предприятиях"
            sheet["E4"] = " В крестьянских/фермерских хозяйствах"
            sheet["F4"] = "  На других предприятиях и у ИП (торговых точках)"
            sheet["G4"] = " Запасы на складах (ОРЦ, ТЛЦ, оптовых реализаторов)"
            sheet["H4"] = "В Стабфондах"
            sheet["I4"] = "  Запасы овощехранилищ"
            sheet["J4"] = "  Запасы фруктохранилищ"

        latest = self.latest_by_region()
        report_dates = [
            datetime.fromisoformat(row["report_date"]).date()
            for row in latest.values()
            if row["report_date"]
        ]
        report_date = max(report_dates) if report_dates else date.today()
        report_date_text = report_date.strftime("%d.%m.%Y")
        sheet["B2"] = f"ЗАПАСЫ ПРОДОВОЛЬСТВЕННЫХ ТОВАРОВ НА {report_date_text} ГОДА"
        sheet["I35"] = f"*данные МИО на {report_date_text}"

        totals = {
            row["product"]: dict(row)
            for row in self.connection.execute(
                """
                SELECT s.product,
                       SUM(s.agricultural_enterprises) agricultural_enterprises,
                       SUM(s.farms) farms,
                       SUM(s.other_enterprises) other_enterprises,
                       SUM(s.warehouses) warehouses,
                       SUM(s.stabilization_direct) stabilization_direct,
                       SUM(s.stabilization_forward) stabilization_forward,
                       SUM(s.stabilization_revolving) stabilization_revolving,
                       SUM(s.vegetable_storage) vegetable_storage,
                       SUM(s.fruit_storage) fruit_storage
                FROM reports r
                JOIN (
                    SELECT region, MAX(report_date) report_date
                    FROM reports
                    GROUP BY region
                ) latest ON latest.region = r.region AND latest.report_date = r.report_date
                JOIN stock_rows s ON s.report_id = r.id
                GROUP BY s.product
                """
            ).fetchall()
        }

        def number(value: float | int | None) -> float | int | None:
            if value is None:
                return None
            value = round(float(value), 6)
            if abs(value) < 0.0000005:
                return None
            return int(value) if float(value).is_integer() else value

        for row_number, display_name, product in SUMMARY_ROWS:
            data = totals.get(product, {})
            sheet.cell(row_number, 2).value = display_name
            sheet.cell(row_number, 3).value = f"=D{row_number}+E{row_number}+F{row_number}+G{row_number}+I{row_number}+H{row_number}"
            sheet.cell(row_number, 4).value = number(data.get("agricultural_enterprises"))
            sheet.cell(row_number, 5).value = number(data.get("farms"))
            sheet.cell(row_number, 6).value = number(data.get("other_enterprises"))
            sheet.cell(row_number, 7).value = number(data.get("warehouses"))
            sheet.cell(row_number, 8).value = number(
                (data.get("stabilization_direct") or 0)
                + (data.get("stabilization_forward") or 0)
                + (data.get("stabilization_revolving") or 0)
            )
            sheet.cell(row_number, 9).value = number(data.get("vegetable_storage"))
            sheet.cell(row_number, 10).value = number(data.get("fruit_storage"))

        for column in range(3, 11):
            letter = sheet.cell(34, column).column_letter
            sheet.cell(34, column).value = f"=SUM({letter}5:{letter}31)"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.save(path)
        return path
