import math
import re
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from .models import Snapshot, SourceFile

MONTHS = {
    "января": 1, "февраля": 2, "марта": 3, "апреля": 4,
    "мая": 5, "июня": 6, "июля": 7, "августа": 8,
    "сентября": 9, "октября": 10, "ноября": 11, "декабря": 12,
}


def normalize(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def key(value: object) -> str:
    return normalize(value).casefold()


def product_key(value: object) -> str:
    value_key = key(value)
    return re.sub(r",\s*(литр|десяток)$", "", value_key)


def _is_number(value: object) -> bool:
    return (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
        and math.isfinite(value)
    )


def extract_date(text: str) -> date | None:
    text = normalize(text).lower()
    iso = re.search(r"(20\d{2})-(\d{1,2})-(\d{1,2})", text)
    if iso:
        year, month, day = map(int, iso.groups())
        return date(year, month, day)
    words = re.search(
        r"(\d{1,2})\s+(" + "|".join(MONTHS) + r")\s+(20\d{2})", text
    )
    if words:
        return date(int(words.group(3)), MONTHS[words.group(2)], int(words.group(1)))
    numeric = re.search(r"(\d{1,2})[./](\d{1,2})[./](20\d{2})", text)
    if numeric:
        day, month, year = map(int, numeric.groups())
        return date(year, month, day)
    return None


def _observation_date(workbook, path: Path) -> date:
    candidates = [path.stem]
    # Сначала заголовки таблиц с датой наблюдения, затем обложка,
    # где также присутствует более поздняя дата публикации.
    for sheet_name in ("1", "5", "Обложка"):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for row in range(1, min(sheet.max_row, 25) + 1):
            for col in range(1, min(sheet.max_column, 5) + 1):
                value = normalize(sheet.cell(row, col).value)
                if value:
                    candidates.append(value)
    for text in candidates:
        parsed = extract_date(text)
        if parsed:
            return parsed
    raise ValueError(f"{path.name}: не удалось определить дату периода.")


def discover_source_files(raw_dir: Path) -> list[SourceFile]:
    if not raw_dir.exists():
        raise FileNotFoundError(f"Папка исходных данных не найдена: {raw_dir}")
    by_date: dict[date, SourceFile] = {}
    paths = sorted(
        path for path in raw_dir.iterdir()
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xls"}
    )
    for path in paths:
        if path.suffix.lower() == ".xls":
            raise RuntimeError(
                f"{path.name}: формат .xls не поддерживается; сохраните файл как .xlsx."
            )
        workbook = load_workbook(path, data_only=True, read_only=True)
        try:
            period = _observation_date(workbook, path)
        finally:
            workbook.close()
        # При дублях одной даты предпочитаем исходное имя БНС, а не кэш агента.
        candidate = SourceFile(period, path)
        existing = by_date.get(period)
        if existing is None or not path.name.startswith("bns_prices_"):
            by_date[period] = candidate
    return sorted(by_date.values(), key=lambda item: item.observation_date, reverse=True)


def parse_snapshot(source: SourceFile) -> Snapshot:
    workbook = load_workbook(source.path, data_only=True, read_only=True)
    try:
        price_sheet = next(
            (
                sheet for sheet in workbook.worksheets
                if "средние цены" in key(sheet.cell(1, 1).value)
            ),
            None,
        )
        index_sheet = next(
            (
                sheet for sheet in workbook.worksheets
                if key(sheet.cell(1, 1).value).startswith(
                    "1. индекс цен на социально-значимые"
                )
            ),
            None,
        )
        if price_sheet is None or index_sheet is None:
            raise ValueError(
                f"{source.path.name}: не найдены листы средних цен или индексов."
            )

        header_row = next(
            row for row in range(1, min(price_sheet.max_row, 15) + 1)
            if sum(
                bool(normalize(price_sheet.cell(row, col).value))
                for col in range(2, price_sheet.max_column + 1)
            ) >= 3
        )
        regions = {
            col: normalize(price_sheet.cell(header_row, col).value)
            for col in range(2, price_sheet.max_column + 1)
            if normalize(price_sheet.cell(header_row, col).value)
        }
        prices: dict[str, dict[str, float]] = {}
        for row in range(header_row + 1, price_sheet.max_row + 1):
            product = normalize(price_sheet.cell(row, 1).value)
            values = {
                region: float(price_sheet.cell(row, col).value)
                for col, region in regions.items()
                if _is_number(price_sheet.cell(row, col).value)
            }
            if product and values:
                prices[product_key(product)] = values

        annual_change: dict[str, float] = {}
        year_change: dict[str, float] = {}
        week_change: dict[str, float] = {}
        for row in range(5, index_sheet.max_row + 1):
            product = normalize(index_sheet.cell(row, 1).value)
            annual_index = index_sheet.cell(row, 2).value
            year_index = index_sheet.cell(row, 3).value
            week_index = index_sheet.cell(row, 5).value
            if not product:
                continue
            if _is_number(annual_index):
                annual_change[product_key(product)] = round(
                    float(annual_index) - 100, 1
                )
            if _is_number(year_index):
                year_change[product_key(product)] = round(float(year_index) - 100, 1)
            if _is_number(week_index):
                week_change[product_key(product)] = round(float(week_index) - 100, 1)
    finally:
        workbook.close()
    if not prices:
        raise ValueError(f"{source.path.name}: средние цены не распознаны.")
    return Snapshot(
        source=source,
        prices=prices,
        annual_change=annual_change,
        year_change=year_change,
        week_change=week_change,
    )
